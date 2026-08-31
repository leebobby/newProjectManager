"""迭代下的需求条目 CRUD + Excel 批量导入。

权限：协作编辑域（见 CLAUDE.md「Write-permission principle」）——日常填报，
登录用户均可读写，与客户面状态的非管理员字段一致。
"""
import io
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import enums
import models
import schemas
from auth import get_current_user
from database import get_db
from op_log import log_op
from notify import dispatch
from routers._lookups import (
    fill_group_fk, fill_user_fk, fill_version_fk, project_name_map, resolve_project_id,
)
from routers._req_dedup import dedup_key, duplicate_message, find_duplicate

router = APIRouter(prefix="/api/iteration-requirements", tags=["iteration-requirements"])


# Excel 导入/模板使用的列定义：(列标题, 模型字段名, 是否必填)
_IMPORT_COLUMNS = [
    ("序号", "seq", False),
    ("需求编号", "req_no", False),
    ("需求超链接", "req_url", False),
    ("需求标题", "title", True),
    ("项目", "_project_name", False),
    ("责任人", "owner", False),
    ("PL组", "owner_group", False),
    ("优先级", "priority", False),
    ("计划交付版本", "planned_version", False),
    ("需求串讲", "progress_walkthrough", False),
    ("反串讲", "progress_reverse", False),
    ("STC设计", "progress_stc", False),
    ("编码", "progress_coding", False),
    ("BBIT", "progress_bbit", False),
    ("转测澄清", "progress_clarify", False),
    ("备注", "remark", False),
]

# 词表统一收口到 enums（见 enums.py）；导入路径仍用集合做快速校验。
_PROGRESS_VALID = set(enums.PROGRESS_STATUSES)
_PRIORITY_VALID = set(enums.PRIORITIES)


def _out(item: models.IterationRequirement, pmap: dict) -> schemas.IterationRequirementOut:
    out = schemas.IterationRequirementOut.model_validate(item)
    out.project_name = pmap.get(item.project_id)
    return out


@router.get("", response_model=List[schemas.IterationRequirementOut])
def list_by_iteration(
    iteration_id: int = Query(..., description="迭代 ID"),
    project_id: Optional[int] = Query(None, description="只看某个项目；不传＝全部（含未指定项目的行）"),
    db: Session = Depends(get_db),
):
    q = (
        db.query(models.IterationRequirement)
        .filter(models.IterationRequirement.iteration_id == iteration_id)
    )
    if project_id is not None:
        q = q.filter(models.IterationRequirement.project_id == project_id)
    items = q.order_by(models.IterationRequirement.seq.asc(),
                       models.IterationRequirement.id.asc()).all()
    pmap = project_name_map(db)
    return [_out(i, pmap) for i in items]


@router.get("/by-version", response_model=List[schemas.IterationRequirementOut])
def list_by_version(
    version_id: int = Query(..., description="迭代版本 ID（按 target_version_id 过滤）"),
    db: Session = Depends(get_db),
):
    """版本管理用：列出"计划交付版本"指向该迭代版本的领域需求。"""
    items = (
        db.query(models.IterationRequirement)
        .filter(models.IterationRequirement.target_version_id == version_id)
        .order_by(models.IterationRequirement.seq.asc(), models.IterationRequirement.id.asc())
        .all()
    )
    pmap = project_name_map(db)
    return [_out(i, pmap) for i in items]


@router.post("", response_model=schemas.IterationRequirementOut)
def create_item(
    payload: schemas.IterationRequirementCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    parent = (
        db.query(models.AnnualIteration)
        .filter(models.AnnualIteration.id == payload.iteration_id)
        .first()
    )
    if not parent:
        raise HTTPException(status_code=404, detail="所属迭代不存在")
    # 同一迭代里已经有这条需求就别再录一条（判重口径见 routers/_req_dedup.py）
    dup = find_duplicate(db, models.IterationRequirement, payload.iteration_id,
                         payload.req_no, payload.title)
    if dup is not None:
        raise HTTPException(status_code=409, detail=duplicate_message(dup))
    # 序号自动取当前最大值+1（如果调用方没传或传 0）
    data = payload.model_dump()
    if not data.get("seq"):
        max_seq = (
            db.query(models.IterationRequirement)
            .filter(models.IterationRequirement.iteration_id == payload.iteration_id)
            .count()
        )
        data["seq"] = max_seq + 1
    fill_user_fk(db, data, "owner", "owner_user_id")
    fill_group_fk(db, data, "owner_group", "group_id")
    fill_version_fk(db, data, "planned_version", "target_version_id")
    item = models.IterationRequirement(**data)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_op(db, action="新增", target="迭代需求", target_id=item.id,
           detail=f"iteration_id={item.iteration_id} project_id={item.project_id} title={item.title}",
           user=current_user, request=request)
    return _out(item, project_name_map(db))


@router.put("/{item_id}", response_model=schemas.IterationRequirementOut)
def update_item(
    item_id: int,
    payload: schemas.IterationRequirementUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    item = (
        db.query(models.IterationRequirement)
        .filter(models.IterationRequirement.id == item_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    if item.version != payload.version:
        raise HTTPException(status_code=409, detail="数据已被他人修改，请刷新后重试")
    changes = payload.model_dump(exclude_unset=True)
    changes.pop("version", None)
    # 编辑也能造出重复（把编号改成另一行的），所以这里同样要判——
    # 用改完之后的值去比，并把自己排除掉
    if "req_no" in changes or "title" in changes:
        dup = find_duplicate(
            db, models.IterationRequirement, item.iteration_id,
            changes.get("req_no", item.req_no), changes.get("title", item.title),
            exclude_id=item.id,
        )
        if dup is not None:
            raise HTTPException(status_code=409, detail=duplicate_message(dup))
    # 字符串改写时同步刷新 FK；若调用方显式传了 FK 字段，尊重它
    fill_user_fk(db, changes, "owner", "owner_user_id")
    fill_group_fk(db, changes, "owner_group", "group_id")
    fill_version_fk(db, changes, "planned_version", "target_version_id")

    old_owner_id = item.owner_user_id

    for k, v in changes.items():
        setattr(item, k, v)
    item.version += 1
    db.commit()
    db.refresh(item)
    log_op(db, action="修改", target="迭代需求", target_id=item.id,
           detail=f"title={item.title} fields={','.join(changes.keys()) or '无'}",
           user=current_user, request=request)

    # 通知策略（大颗粒）：只在「被指派为负责人」时提醒；逐条进展字段变更不再发通知，
    # 避免填报 70-80 条需求时跑马灯被刷屏（版本类通知另行保留细颗粒）。
    link = f"/iterations/{item.iteration_id}"
    if item.owner_user_id and item.owner_user_id != old_owner_id:
        dispatch(
            db, kind="assignment",
            title=f"你被指派为需求负责人：{item.title or ''}",
            body="", link=link,
            source_type="iteration_requirement", source_id=item.id,
            actor=current_user, recipient_ids=[item.owner_user_id], extra_subs=False,
        )
    return _out(item, project_name_map(db))


@router.delete("/{item_id}")
def delete_item(
    item_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    item = (
        db.query(models.IterationRequirement)
        .filter(models.IterationRequirement.id == item_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    snapshot = f"title={item.title}"
    db.delete(item)
    db.commit()
    log_op(db, action="删除", target="迭代需求", target_id=item_id,
           detail=snapshot, user=current_user, request=request)
    return {"ok": True}


@router.get("/import-template.xlsx")
def download_import_template():
    """下载批量导入模板（含表头行 + 一行示例）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    from xlsx_io import beautify, style_header

    wb = Workbook()
    ws = wb.active
    ws.title = "需求清单"

    headers = [c[0] for c in _IMPORT_COLUMNS]
    style_header(ws, headers)

    # 示例行
    example = [
        1, "REQ-2026-001", "https://example.com/req/2026-001", "示例需求标题",
        "YLS3000", "张三", "AFK", "P1", "v0.6.0",
        "已完成", "已完成", "进行中", "进行中", "未开始", "未开始",
        "需求范围已变更，原范围……",
    ]
    ws.append(example)

    # 列宽（与 _IMPORT_COLUMNS 一一对应）
    widths = [6, 16, 26, 32, 14, 10, 10, 8, 14, 10, 10, 10, 10, 10, 12, 30]
    for i, w in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    beautify(ws, last_row=ws.max_row)

    # 提示行（合并单元格）
    tip_row = ws.max_row + 2
    ws.cell(row=tip_row, column=1, value=(
        "提示：进展列填写「未开始/进行中/已完成/已延期/已变更/不涉及」之一；优先级填 P0/P1/P2/P3；"
        "「项目」要与系统里的项目名完全一致，对不上会留空（导入后可在页面上补选，不会报错）；"
        "序号留空将自动按现有最大序号顺序累加；删除示例行后再导入。"
    )).font = Font(italic=True, color="909399")
    ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "iteration-requirements-template.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import")
async def import_from_excel(
    request: Request,
    iteration_id: int = Query(..., description="目标迭代 ID"),
    file: UploadFile = File(..., description="xlsx 文件"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """从 xlsx 批量导入需求到指定迭代。返回 {created, errors}。"""
    parent = (
        db.query(models.AnnualIteration)
        .filter(models.AnnualIteration.id == iteration_id)
        .first()
    )
    if not parent:
        raise HTTPException(status_code=404, detail="所属迭代不存在")

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    from openpyxl import load_workbook

    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"读取 xlsx 失败：{exc}")

    ws = wb.active
    if ws.max_row < 2:
        raise HTTPException(status_code=400, detail="文件为空")

    # 用首行表头匹配字段：列标题 -> 模型字段
    header_row = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col_map = {label: field for label, field, _ in _IMPORT_COLUMNS}
    field_idx = {}
    for idx, label in enumerate(header_row):
        if label in col_map:
            field_idx[col_map[label]] = idx

    if "title" not in field_idx:
        raise HTTPException(status_code=400, detail="模板缺少必填列「需求标题」")

    current_max = (
        db.query(models.IterationRequirement)
        .filter(models.IterationRequirement.iteration_id == iteration_id)
        .count()
    )

    created = 0
    skipped = 0
    errors: List[str] = []
    pending = []
    # 判重用的已有键：一次查完，别在循环里逐行查库（一份表格几百行）
    existing_keys = {}
    for row in (db.query(models.IterationRequirement)
                .filter(models.IterationRequirement.iteration_id == iteration_id).all()):
        k = dedup_key(row.req_no, row.title)
        if k is not None:
            existing_keys.setdefault(k, row)
    seen_in_file = {}

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过完全空行 / 提示行（首列以"提示"开头）
        if not row or all(v in (None, "") for v in row):
            continue
        first_val = row[0] if len(row) > 0 else None
        if isinstance(first_val, str) and first_val.strip().startswith("提示"):
            continue

        data = {}
        for field, idx in field_idx.items():
            if idx >= len(row):
                continue
            v = row[idx]
            if v is None:
                continue
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            data[field] = v if isinstance(v, (int, float)) else str(v).strip()

        title = data.get("title")
        if not title:
            errors.append(f"第 {r_idx} 行：缺少需求标题，已跳过")
            continue

        # 重复的行**跳过而不是报错**：一次导入里混着几条已录过的很正常（补录时
        # 常常把整张表重新导一遍），当成错误会让人以为整份表格有问题。但要如实
        # 报出来——只跳不报的表现是「导入 80 条只进了 60 条」，而没人说得清少的是哪些。
        key = dedup_key(data.get("req_no"), title)
        if key is not None:
            if key in existing_keys:
                skipped += 1
                errors.append(f"第 {r_idx} 行：{duplicate_message(existing_keys[key], '本迭代里已有')}"
                              "本行已跳过")
                continue
            if key in seen_in_file:
                skipped += 1
                errors.append(f"第 {r_idx} 行：与本文件第 {seen_in_file[key]} 行重复，已跳过")
                continue
            seen_in_file[key] = r_idx

        # 校验枚举
        progress_fields = [
            "progress_walkthrough", "progress_reverse", "progress_stc",
            "progress_coding", "progress_bbit", "progress_clarify",
        ]
        bad_progress = False
        for pf in progress_fields:
            if pf in data and data[pf] not in _PROGRESS_VALID:
                errors.append(f"第 {r_idx} 行：进展列「{pf}」值「{data[pf]}」非法，已跳过整行")
                bad_progress = True
                break
        if bad_progress:
            continue

        priority = data.get("priority")
        if priority and priority not in _PRIORITY_VALID:
            errors.append(f"第 {r_idx} 行：优先级「{priority}」非法，已跳过")
            continue

        seq = data.get("seq")
        if not isinstance(seq, int) or seq <= 0:
            current_max += 1
            data["seq"] = current_max

        data["iteration_id"] = iteration_id
        pending.append(data)

    for d in pending:
        fill_user_fk(db, d, "owner", "owner_user_id")
        fill_group_fk(db, d, "owner_group", "group_id")
        fill_version_fk(db, d, "planned_version", "target_version_id")
        # 「项目」列不是模型字段，只是反查 FK 的入参，落库前摘掉。
        # 反查不中留空而不报错，交给页面事后补选（见 CLAUDE.md「主数据与 FK 反查」）。
        d["project_id"] = resolve_project_id(db, d.pop("_project_name", None))
        item = models.IterationRequirement(**d)
        db.add(item)
        created += 1
    db.commit()

    log_op(db, action="导入", target="迭代需求", target_id=iteration_id,
           detail=f"created={created} skipped={skipped} errors={len(errors)} "
                  f"file={file.filename or ''}",
           user=current_user, request=request)
    return {"created": created, "skipped": skipped, "errors": errors}
