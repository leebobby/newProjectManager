"""统一的 Excel 导出美化。

所有业务导出（清单类平表）共用一套外观，配色取自 [brand.py](brand.py)，
与 PPT 导出**同一套**（见 CLAUDE.md「导出与上传」的「红定位、蓝分层、饱和色点灯」）：
- 表头：浅蓝灰 #D9E2F3 填充 + 黑字加粗 + 居中 + 换行
- 数据区：细边框、斑马纹（几乎看不见的浅蓝灰）、自动换行、垂直居中
- 状态点灯：值恰好等于「已完成 / 进行中 / 已延期 / 已变更」的格子上红黄绿灰底色
- 冻结表头（已设置 freeze_panes 的沿用，不覆盖）
- 列宽：按内容显示宽度自适应（中文按 2 计），设了显式列宽的列不动

用法（在数据行全部写完之后、追加"提示行"之前调用）：

    from xlsx_io import style_header, beautify
    style_header(ws, headers)          # 写入并美化第 1 行表头
    ... ws.append(数据行) ...
    beautify(ws)                       # 边框/斑马/列宽/冻结一次成型

提示行等附注内容在 beautify() 之后再追加，不参与表格样式。
"""
from typing import Iterable, Optional, Sequence

import brand
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 表头的品牌色。曾经是品牌蓝 #4073BA + 白字，与 PPT 的红底白字各走各的；
# 现在统一收口到 brand.py，**不要在这里另写颜色字面量**。
BRAND = brand.HEADER_BG

_HEAD_FONT = Font(bold=True, color=brand.HEADER_TEXT, size=11)
_HEAD_FILL = PatternFill("solid", fgColor=brand.HEADER_BG)
_HEAD_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_THIN = Side(style="thin", color=brand.BORDER)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ZEBRA_FILL = PatternFill("solid", fgColor=brand.ZEBRA)
_DATA_ALIGN = Alignment(vertical="center", wrap_text=True)
_DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_FONT = Font(size=11, color=brand.TEXT)

# 状态点灯的 Font/PatternFill 预先建好：openpyxl 的样式对象可以共享，
# 每格新建一个的话，几万格的导出光建样式就要好几秒。
_STATUS_FILL_OBJ = {k: PatternFill("solid", fgColor=v) for k, v in brand.STATUS_FILLS.items()}
_STATUS_FONT_OBJ = {k: Font(size=11, color=brand.STATUS_ON_FILL_TEXT, bold=True)
                    for k in brand.STATUS_FILLS}
_STATUS_FONT_OBJ.update({k: Font(size=11, color=v) for k, v in brand.STATUS_TEXT.items()})


def style_header(ws, headers: Sequence[str], row: int = 1):
    """在第 row 行写入表头并套用品牌样式。已有同名局部实现的路由可直接换用这里。"""
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row, j, h)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.alignment = _HEAD_ALIGN
        c.border = _BORDER
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 24)


def _disp_w(v) -> int:
    """显示宽度：全角字符按 2 计；多行取最长一行。"""
    s = "" if v is None else str(v)
    best = 0
    for line in s.split("\n"):
        w = sum(2 if ord(ch) > 0x2E7F else 1 for ch in line)
        best = max(best, w)
    return best


def beautify(
    ws,
    header_rows: int = 1,
    last_row: Optional[int] = None,
    center_cols: Iterable[int] = (),
    max_width: int = 50,
    min_width: int = 7,
    zebra: bool = True,
    light_status: bool = True,
):
    """给已写完数据的工作表补边框/斑马纹/换行/列宽/冻结表头。

    - header_rows: 表头占几行（分组表头传 2）
    - last_row: 数据区最后一行（默认 ws.max_row；先 beautify 再追加提示行可自然排除附注）
    - center_cols: 需要水平居中的列号集合（1 起）；其余列左对齐
    - 列宽只在该列没有显式宽度时设置，模板里手工调过的列宽保持不变
    - light_status: 值**恰好等于**状态词的格子上底色（绿/黄/红/灰）。
      精确匹配，不做包含匹配——"已完成三个模块联调"整格染绿是错的。
      个别表不想要（比如状态词只是普通文本）传 False 关掉。
    """
    end = last_row or ws.max_row
    if end < 1:
        return
    n_cols = ws.max_column
    center = set(center_cols)

    widths = [0] * (n_cols + 1)
    for row in ws.iter_rows(min_row=1, max_row=end, max_col=n_cols):
        for cell in row:
            j = cell.column
            widths[j] = max(widths[j], _disp_w(cell.value))
            if cell.row <= header_rows:
                continue  # 表头由 style_header 负责
            cell.border = _BORDER
            key = str(cell.value).strip() if cell.value is not None else ""
            lit = _STATUS_FILL_OBJ.get(key) if light_status else None
            cell.font = (_STATUS_FONT_OBJ.get(key) or _DATA_FONT) if light_status else _DATA_FONT
            # 点灯的格子一律居中：状态词都很短，靠左时在宽列里飘着找不到
            cell.alignment = _DATA_ALIGN_CENTER if (j in center or lit) else _DATA_ALIGN
            if lit is not None:
                cell.fill = lit
            elif zebra and (cell.row - header_rows) % 2 == 0:
                cell.fill = _ZEBRA_FILL

    from openpyxl.utils import get_column_letter
    for j in range(1, n_cols + 1):
        letter = get_column_letter(j)
        dim = ws.column_dimensions[letter]
        if dim.width:  # 显式设置过的列宽尊重原值
            continue
        dim.width = min(max_width, max(min_width, widths[j] + 3))

    if not ws.freeze_panes:
        ws.freeze_panes = f"A{header_rows + 1}"
