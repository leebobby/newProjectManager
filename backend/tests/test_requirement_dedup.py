"""需求判重：同一迭代里同一条需求只该录一次。

历史上手工新增和 Excel 导入都不判重，于是同一份表格导两遍就是整份翻倍、
补录时又手敲一条就多一行。重复的需求在度量里是实打实的分母——完成度被摊薄、
按项目/领域的条数偏大，而每一行单独看都合法，没人会当 bug 报上来。

口径（`routers/_req_dedup.py`，领域需求与产品需求共用一份实现）：
- 范围是**一个迭代**，跨迭代不拦（同一条需求下个月接着排是正常的）；
- 有需求编号按编号判，没编号按标题判；
- 比较前把空白去掉再转小写（Excel 粘出来的编号常带首尾空格）。
"""
import datetime
import io

import pytest


@pytest.fixture(scope="module")
def iters(client, admin_headers):
    rows = client.get("/api/annual-iterations", headers=admin_headers,
                      params={"year": datetime.date.today().year}).json()
    assert len(rows) >= 2, "年度迭代应自动生成 12 条"
    return rows[0]["id"], rows[1]["id"]


def _post(client, headers, path, iteration_id, **kw):
    body = {"iteration_id": iteration_id}
    body.update(kw)
    return client.post(path, headers=headers, json=body)


DOMAIN = "/api/iteration-requirements"
PRODUCT = "/api/iteration-product-requirements"


# ── 手工新增 ────────────────────────────────────────────────────────────────

def test_same_req_no_in_one_iteration_is_rejected(client, admin_headers, iters):
    it, _ = iters
    ok = _post(client, admin_headers, DOMAIN, it, req_no="REQ-1001", title="登录超时")
    assert ok.status_code == 200, ok.text

    dup = _post(client, admin_headers, DOMAIN, it, req_no="REQ-1001", title="标题写得不一样")
    assert dup.status_code == 409
    detail = dup.json()["detail"]
    assert "REQ-1001" in detail and "登录超时" in detail, "提示要指到具体是哪一行"


def test_req_no_compare_ignores_spaces_and_case(client, admin_headers, iters):
    """Excel 里粘出来的编号常带首尾空格或全角空格，肉眼看不出差别。"""
    it, _ = iters
    assert _post(client, admin_headers, DOMAIN, it,
                 req_no="REQ-2001", title="甲").status_code == 200
    for variant in (" REQ-2001 ", "REQ-2001　", "req-2001"):
        r = _post(client, admin_headers, DOMAIN, it, req_no=variant, title="乙")
        assert r.status_code == 409, f"{variant!r} 应判为重复"


def test_no_req_no_falls_back_to_title(client, admin_headers, iters):
    """编号是选填的，只按编号判等于"不填编号就能重复录"。"""
    it, _ = iters
    assert _post(client, admin_headers, DOMAIN, it, title="报表导出很慢").status_code == 200
    assert _post(client, admin_headers, DOMAIN, it,
                 title=" 报表导出很慢 ").status_code == 409


def test_same_req_no_in_another_iteration_is_fine(client, admin_headers, iters):
    """同一条需求本轮没做完、下个月接着排是正常的，跨迭代拦住会逼着人改标题绕过去。"""
    it_a, it_b = iters
    assert _post(client, admin_headers, DOMAIN, it_a,
                 req_no="REQ-3001", title="跨迭代").status_code == 200
    assert _post(client, admin_headers, DOMAIN, it_b,
                 req_no="REQ-3001", title="跨迭代").status_code == 200


def test_product_requirements_share_the_same_rule(client, admin_headers, iters):
    """两张表共用一份实现——两处各写一份的表现是一个 Tab 拦住了、另一个没拦。"""
    it, _ = iters
    assert _post(client, admin_headers, PRODUCT, it,
                 req_no="P-1", title="产品甲").status_code == 200
    assert _post(client, admin_headers, PRODUCT, it,
                 req_no="P-1", title="产品乙").status_code == 409


# ── 编辑也能造出重复 ────────────────────────────────────────────────────────

def test_editing_a_req_no_into_another_row_is_rejected(client, admin_headers, iters):
    it, _ = iters
    a = _post(client, admin_headers, DOMAIN, it, req_no="REQ-4001", title="甲").json()
    b = _post(client, admin_headers, DOMAIN, it, req_no="REQ-4002", title="乙").json()

    r = client.put(f"{DOMAIN}/{b['id']}", headers=admin_headers,
                   json={"req_no": "REQ-4001", "version": b["version"]})
    assert r.status_code == 409
    assert "REQ-4001" in r.json()["detail"]

    # 改别的字段不受影响，而且一行不算与自己重复
    r = client.put(f"{DOMAIN}/{b['id']}", headers=admin_headers,
                   json={"req_no": "REQ-4002", "title": "乙改名", "version": b["version"]})
    assert r.status_code == 200, r.text
    assert a["req_no"] == "REQ-4001"


# ── Excel 导入 ──────────────────────────────────────────────────────────────

def _make_xlsx(client, headers, path, rows):
    """按真实模板的表头拼一份 xlsx，避免测试里另写一套列名。"""
    from openpyxl import load_workbook
    resp = client.get(f"{path}/import-template.xlsx", headers=headers)
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    ws.delete_rows(2, ws.max_row)   # 去掉模板自带的示例行/提示行
    for r in rows:
        ws.append([r.get(h, "") for h in header])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, headers, path, iteration_id, rows):
    blob = _make_xlsx(client, headers, path, rows)
    return client.post(f"{path}/import", headers=headers,
                       params={"iteration_id": iteration_id},
                       files={"file": ("req.xlsx", blob,
                                       "application/vnd.openxmlformats-officedocument."
                                       "spreadsheetml.sheet")})


def test_importing_the_same_file_twice_does_not_double_everything(client, admin_headers, iters):
    it, _ = iters
    rows = [{"需求编号": "IMP-1", "需求标题": "导入甲"},
            {"需求编号": "IMP-2", "需求标题": "导入乙"}]

    first = _upload(client, admin_headers, DOMAIN, it, rows).json()
    assert first["created"] == 2 and first["skipped"] == 0

    second = _upload(client, admin_headers, DOMAIN, it, rows).json()
    assert second["created"] == 0, "整份表格重导不该再进一遍"
    assert second["skipped"] == 2
    # 只跳不报的表现是「导入 80 条只进了 60 条」，没人说得清少的是哪些
    assert len(second["errors"]) == 2
    assert all("已跳过" in e for e in second["errors"])


def test_duplicates_inside_one_file_are_skipped_too(client, admin_headers, iters):
    it, _ = iters
    rows = [{"需求编号": "IMP-9", "需求标题": "文件内甲"},
            {"需求编号": " imp-9 ", "需求标题": "文件内甲的另一种写法"}]
    d = _upload(client, admin_headers, DOMAIN, it, rows).json()
    assert d["created"] == 1 and d["skipped"] == 1
    assert "第 2 行" in d["errors"][0], "要指出跟本文件的哪一行撞了"


def test_import_still_creates_the_new_rows(client, admin_headers, iters):
    """重复的跳过，不重复的照进——别把整份表格判成失败。"""
    it, _ = iters
    _upload(client, admin_headers, DOMAIN, it, [{"需求编号": "IMP-20", "需求标题": "已存在"}])
    d = _upload(client, admin_headers, DOMAIN, it, [
        {"需求编号": "IMP-20", "需求标题": "已存在"},
        {"需求编号": "IMP-21", "需求标题": "新的"},
    ]).json()
    assert d["created"] == 1 and d["skipped"] == 1


# ── 跨迭代：不拦，但要报出来 ────────────────────────────────────────────────

def test_cross_iteration_duplicate_is_reported_not_blocked(client, admin_headers, iters):
    """本轮没做完、下个月接着排是正常的，所以不拦；但也可能是重复录了，所以要报。"""
    it_a, it_b = iters
    _post(client, admin_headers, DOMAIN, it_a, req_no="X-1", title="跨迭代提示")
    ok = _post(client, admin_headers, DOMAIN, it_b, req_no="X-1", title="跨迭代提示")
    assert ok.status_code == 200, "跨迭代不该拦"

    d = client.get(f"{DOMAIN}/duplicates", headers=admin_headers,
                   params={"iteration_id": it_b}).json()
    g = next(g for g in d["groups"] if any(r["req_no"] == "X-1" for r in g["rows"]))
    assert g["kind"] == "no"
    assert len(g["rows"]) == 1, "本迭代里只有一条"
    assert [o["iteration_id"] for o in g["others"]] == [it_a]
    assert g["others"][0]["iteration_label"], "要写清楚是哪个迭代，光给 id 没人查得动"
    assert d["cross_iteration"] >= 1


def test_duplicates_scan_also_surfaces_legacy_same_iteration_rows(client, admin_headers, iters):
    """判重是后加的，存量数据里的重复不会自己消失——扫描要能把它们捞出来。"""
    import models
    from database import SessionLocal

    it, _ = iters
    a = _post(client, admin_headers, DOMAIN, it, req_no="LEGACY-1", title="老重复").json()
    # 绕过接口直接塞一条，模拟判重上线之前就已经在库里的行
    db = SessionLocal()
    db.add(models.IterationRequirement(iteration_id=it, seq=999,
                                       req_no="LEGACY-1", title="老重复（另一条）"))
    db.commit()
    db.close()

    d = client.get(f"{DOMAIN}/duplicates", headers=admin_headers,
                   params={"iteration_id": it}).json()
    g = next(g for g in d["groups"] if any(r["req_no"] == "LEGACY-1" for r in g["rows"]))
    assert len(g["rows"]) == 2, "同一迭代里的两条都要列出来"
    assert a["id"] in [r["id"] for r in g["rows"]]
    assert d["same_iteration"] >= 1


def test_clean_iteration_reports_nothing(client, admin_headers, iters):
    """没有重复时不要报一堆——提示条只在真有问题时出现。"""
    _, it_b = iters
    d = client.get(f"{DOMAIN}/duplicates", headers=admin_headers,
                   params={"iteration_id": it_b}).json()
    for g in d["groups"]:
        assert len(g["rows"]) > 1 or g["others"], "列出来的每一组都得真有重复"


def test_import_flags_cross_iteration_rows_but_still_imports_them(client, admin_headers, iters):
    it_a, it_b = iters
    _post(client, admin_headers, DOMAIN, it_a, req_no="XIMP-1", title="上个月就有")
    d = _upload(client, admin_headers, DOMAIN, it_b,
                [{"需求编号": "XIMP-1", "需求标题": "上个月就有"}]).json()
    assert d["created"] == 1 and d["skipped"] == 0, "跨迭代照进"
    assert d["cross_iteration"] == 1
    assert any("请确认是否重复" in e for e in d["errors"]), "进了也要说一声"


def test_product_tab_has_its_own_scan(client, admin_headers, iters):
    it_a, it_b = iters
    _post(client, admin_headers, PRODUCT, it_a, req_no="PX-1", title="产品跨迭代")
    _post(client, admin_headers, PRODUCT, it_b, req_no="PX-1", title="产品跨迭代")
    d = client.get(f"{PRODUCT}/duplicates", headers=admin_headers,
                   params={"iteration_id": it_b}).json()
    assert d["cross_iteration"] >= 1
