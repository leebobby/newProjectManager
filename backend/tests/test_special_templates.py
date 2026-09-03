"""专项版式模板：套用语义、幂等、只增不删，以及导出/周报是否跟着版式走。

重点不是 CRUD 能不能跑，而是三条容易回归的约定：
1. 套模板**不删数据**——已填的行、模板没提到的分段都要留着；
2. 重复套同一模板是幂等的（按 tkey 认领，不会越套越多分段）；
3. 分段改名/停用后，Excel 与周报必须跟着变，否则页面和汇报口径又会分叉。
"""
import io
import json

import pytest


def _content(client, headers, sid):
    return client.get(f"/api/specials/{sid}", headers=headers).json()["content"]


def _cfg(content):
    return json.loads(content["section_config_json"])


def _blocks(content):
    return json.loads(content["extra_grids_json"])


def _by_tkey(blocks, tkey):
    """按 tkey 取分段，而不是按下标。

    模板迟早会加分段（0009 就往里加了三段），按下标取的断言那时会集体误报，
    而且报的还是「数据被覆盖」这种看起来很吓人的错。
    """
    b = next((x for x in blocks if x.get("tkey") == tkey), None)
    assert b is not None, f"模板里应有 tkey={tkey} 的分段：{[x.get('tkey') for x in blocks]}"
    return b


@pytest.fixture(scope="module")
def solution_tpl(client, admin_headers):
    """seed 注入的「解决方案专项」模板。"""
    rows = client.get("/api/special-templates", headers=admin_headers).json()
    tpl = next((t for t in rows if t["name"] == "解决方案专项"), None)
    assert tpl, f"seed 应注入解决方案专项模板，实际：{[t['name'] for t in rows]}"
    return tpl


def test_seed_templates_present(client, admin_headers):
    rows = client.get("/api/special-templates", headers=admin_headers).json()
    names = {t["name"] for t in rows}
    assert {"标准专项", "解决方案专项"} <= names


def test_builtin_sections_endpoint(client, admin_headers):
    data = client.get("/api/special-templates/sections", headers=admin_headers).json()
    keys = [s["key"] for s in data["sections"]]
    # 顺序即默认显示顺序，须与前端 FIXED_KEYS 一致
    assert keys == ["goal", "plan", "progress", "help", "panorama",
                    "risks", "tasks", "formation"]
    assert "light" in data["col_types"], "点灯列格式应对前端可见"


def test_create_with_template_applies_layout(client, admin_headers, solution_tpl):
    resp = client.post("/api/specials", headers=admin_headers, json={
        "name": "某解决方案专项", "kind": "special", "owner": "admin",
        "template_id": solution_tpl["id"],
    })
    assert resp.status_code == 200, resp.text
    sid = resp.json()["id"]
    content = _content(client, admin_headers, sid)

    cfg = _cfg(content)
    assert cfg["template_name"] == "解决方案专项"
    assert cfg["sections"]["goal"]["title"] == "解决方案专项目标"
    assert cfg["sections"]["progress"]["title"] == "整体进展和关键风险"
    assert cfg["sections"]["risks"]["title"] == "关键问题跟踪"
    # 这类专项不按里程碑/阵型汇报
    for k in ("plan", "panorama", "tasks", "formation"):
        assert cfg["sections"][k]["enabled"] is False, k

    blocks = _blocks(content)
    assert [b["title"] for b in blocks] == [
        "解决方案验收标准", "总体测试策略", "专项测试计划",
        "测试详细进展和点灯", "战场关键风险评估", "需求转测和规范度评估"]
    # 文本框分段不该带表格的那套列定义
    strategy = _by_tkey(blocks, "test-strategy")
    assert strategy["kind"] == "text" and strategy["html"] == ""
    assert "headers" not in strategy
    lights = _by_tkey(blocks, "test-lights")
    assert lights["colTypes"][-1] == "light"
    # 点灯列没配候选项时自动给一份红黄绿
    assert lights["colOptions"][-1] == ["绿", "黄", "红"]
    assert len(lights["rows"]) == 3 and len(lights["rows"][0]) == 5

    order = json.loads(content["section_order_json"])
    gid = {b["tkey"]: f"grid:{b['gid']}" for b in blocks}
    assert order[:10] == [
        "goal", "progress",
        gid["acceptance"], gid["test-strategy"], gid["test-plan"],
        gid["test-lights"], gid["battlefield-risk"],
        "risks", gid["req-handover"], "help"]
    # 停用的内置分段仍留在顺序里（只是不显示），不会凭空消失
    assert set(order[10:]) == {"plan", "panorama", "tasks", "formation"}


def test_reapply_is_idempotent_and_keeps_data(client, admin_headers, solution_tpl):
    """重复套用不该新增分段，也不该动已经填进去的行。"""
    sid = client.post("/api/specials", headers=admin_headers, json={
        "name": "复用模板的专项", "kind": "special",
        "template_id": solution_tpl["id"],
    }).json()["id"]

    content = _content(client, admin_headers, sid)
    blocks = _blocks(content)
    _by_tkey(blocks, "test-lights")["rows"][0][0] = {
        "text": "手填的测试项", "align": "left", "color": "", "bold": False}
    # 再挂一个模板里没有的自定义分段，验证套用不会把它删掉
    blocks.append({"gid": "usergid01", "kind": "text",
                   "title": "自己加的文本框", "html": "<p>保留我</p>"})
    resp = client.put(f"/api/specials/{sid}/content", headers=admin_headers, json={
        "version": content["version"],
        "extra_grids_json": json.dumps(blocks, ensure_ascii=False),
    })
    assert resp.status_code == 200, resp.text
    version = resp.json()["version"]

    resp = client.post(f"/api/specials/{sid}/apply-template", headers=admin_headers,
                       json={"template_id": solution_tpl["id"], "version": version})
    assert resp.status_code == 200, resp.text
    after = _blocks(resp.json())

    assert len(after) == 7, "模板 6 段 + 用户自加 1 段，重复套用不应新增"
    assert _by_tkey(after, "test-lights")["rows"][0][0]["text"] == "手填的测试项", \
        "已填数据被覆盖了"
    assert any(b["title"] == "自己加的文本框" for b in after), "模板外的分段被删了"


def test_apply_template_rejects_stale_version(client, admin_headers, solution_tpl):
    sid = client.post("/api/specials", headers=admin_headers, json={
        "name": "版本冲突验证", "kind": "special",
    }).json()["id"]
    resp = client.post(f"/api/specials/{sid}/apply-template", headers=admin_headers,
                       json={"template_id": solution_tpl["id"], "version": 999})
    assert resp.status_code == 409


def test_invalid_layout_rejected(client, admin_headers):
    bad = [
        ({"order": ["不存在的分段"], "config": {}, "blocks": []}, "未知 order key"),
        ({"order": [], "config": {"nope": {}}, "blocks": []}, "未知内置分段"),
        ({"order": [], "config": {}, "blocks": [{"kind": "grid"}]}, "缺 tkey"),
        ({"order": [], "config": {}, "blocks": [
            {"tkey": "a", "kind": "grid", "colTypes": ["bogus"]}]}, "非法列格式"),
    ]
    for layout, why in bad:
        resp = client.post("/api/special-templates", headers=admin_headers, json={
            "name": f"非法模板-{why}",
            "layout_json": json.dumps(layout, ensure_ascii=False),
        })
        assert resp.status_code == 400, f"{why} 应被拒：{resp.text}"


@pytest.fixture(scope="module")
def normal_headers(client, admin_headers):
    """普通登录用户，用来验证权限分档而不是只测 admin 路径。"""
    client.post("/api/users", headers=admin_headers, json={
        "username": "tpl_tester", "full_name": "分段测试员",
        "password": "test1234", "role": "normal", "can_login": True,
    })
    resp = client.post("/api/auth/login",
                       json={"username": "tpl_tester", "password": "test1234"})
    assert resp.status_code == 200, resp.text
    return {"Authorization": f"Bearer {resp.json()['access_token']}"}


def test_permission_tiers(client, admin_headers, normal_headers, solution_tpl):
    """分段改名/停用是协作编辑（登录用户），改模板与套模板是配置（仅 admin）。"""
    sid = client.post("/api/specials", headers=admin_headers, json={
        "name": "权限分档验证", "kind": "special",
    }).json()["id"]
    content = _content(client, normal_headers, sid)

    # 普通用户可以改本专项的分段标题与启停（走 PUT /content 的乐观锁）
    cfg = {"sections": {"goal": {"title": "我改的标题", "enabled": True},
                        "formation": {"title": "", "enabled": False}}}
    resp = client.put(f"/api/specials/{sid}/content", headers=normal_headers, json={
        "version": content["version"],
        "section_config_json": json.dumps(cfg, ensure_ascii=False),
    })
    assert resp.status_code == 200, resp.text
    body = client.get(f"/api/specials/{sid}/report-draft",
                      headers=normal_headers).json()["body"]
    assert "我改的标题" not in body, "目标是空的，空段不该只因改了标题就出现"

    client.put(f"/api/specials/{sid}/content", headers=normal_headers, json={
        "version": resp.json()["version"], "goal": "<p>有内容了</p>",
    })
    body = client.get(f"/api/specials/{sid}/report-draft",
                      headers=normal_headers).json()["body"]
    assert "一、我改的标题" in body

    # 但套模板 / 改模板是 admin 的事
    ver = _content(client, normal_headers, sid)["version"]
    resp = client.post(f"/api/specials/{sid}/apply-template", headers=normal_headers,
                       json={"template_id": solution_tpl["id"], "version": ver})
    assert resp.status_code == 403, resp.text
    resp = client.post("/api/special-templates", headers=normal_headers,
                       json={"name": "普通用户建的模板"})
    assert resp.status_code == 403
    resp = client.delete(f"/api/special-templates/{solution_tpl['id']}",
                         headers=normal_headers)
    assert resp.status_code == 403
    # 读是开放的：建专项对话框和详情页都要用
    assert client.get("/api/special-templates", headers=normal_headers).status_code == 200


def test_export_and_report_follow_layout(client, admin_headers, solution_tpl):
    """分段改名/停用后，Excel 与周报要跟着变。"""
    sid = client.post("/api/specials", headers=admin_headers, json={
        "name": "导出口径验证", "kind": "special",
        "template_id": solution_tpl["id"],
    }).json()["id"]
    content = _content(client, admin_headers, sid)
    blocks = _blocks(content)
    # 点灯表填一行：模板预留的空行是没有内容的，空段本就不该进周报
    # （只填这一张，前面几段整表空白，正好验证空段不占章节号）
    _by_tkey(blocks, "test-lights")["rows"][0] = [
        {"text": t, "align": "left", "color": "", "bold": False}
        for t in ("冒烟测试", "张三", "2026-08-20", "全部通过", "绿")
    ]
    client.put(f"/api/specials/{sid}/content", headers=admin_headers, json={
        "version": content["version"],
        "goal": "<p>目标正文</p>",
        "progress_summary": "<p>进展正文</p>",
        "extra_grids_json": json.dumps(blocks, ensure_ascii=False),
    })

    draft = client.get(f"/api/specials/{sid}/report-draft", headers=admin_headers).json()
    body = draft["body"]
    assert "解决方案专项目标" in body, "周报标题没跟随分段配置"
    assert "整体进展和关键风险" in body
    assert "专项计划" not in body, "停用的分段不该出现在周报里"
    assert "专项阵型" not in body
    # 自定义分段进周报了，且只进填过内容的那张表
    assert "三、测试详细进展和点灯" in body, "自定义分段没进周报"
    assert "冒烟测试" in body
    assert "战场关键风险评估" not in body, "整表空白的自定义分段不该占一节"
    # 编号按实际启用且有内容的分段连续排，不是写死的一~六
    assert "四、求助" not in body and "四、" not in body

    resp = client.post(f"/api/specials/{sid}/report.eml", headers=admin_headers, json={})
    assert resp.status_code == 200
    eml = resp.content.decode("utf-8", "replace")
    assert "测试详细进展和点灯" in eml
    assert "#F0F9EB" in eml or "F0F9EB" in eml, "点灯「绿」应在 HTML 里着色"

    resp = client.get(f"/api/specials/{sid}/export.xlsx", headers=admin_headers)
    assert resp.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    # **一个专项只导出一张表**：自定义表格内联在对应章节下，不再另开页签
    assert wb.sheetnames == ["专项"], wb.sheetnames
    ws = wb["专项"]
    heads = [str(c[0].value) for c in ws.iter_rows(max_col=1) if c[0].value]
    assert any(str(c.value) == "冒烟测试" for row in ws.iter_rows() for c in row), \
        "自定义表格的内容应该就在主表里"
    assert "一、解决方案专项目标" in heads
    assert "二、整体进展和关键风险" in heads
    assert "三、测试详细进展和点灯" in heads
    assert not any("专项计划" in h or "专项阵型" in h for h in heads), \
        "停用的分段不该出现在 Excel 里"
