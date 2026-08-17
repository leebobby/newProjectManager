"""硬件问题清零：CRUD、机台汇总（不涉及剔除）、自定义列、导入导出往返。"""
import io

import openpyxl
import pytest


@pytest.fixture()
def hw_config(monkeypatch):
    """固定 config，避免依赖真实 config.json。"""
    import routers.hardware_issues as hw
    cfg = {
        "hw_issue_sources": ["来料不良", "设计缺陷"],
        "hw_machine_cell_options": ["已清零", "未清零", "不涉及"],
        "hw_machine_cell_na_options": ["不涉及"],
        "hw_extra_columns": [
            {"key": "factory", "label": "负责工厂", "type": "text", "after": "__start__"},
            {"key": "grade", "label": "等级", "type": "select", "options": ["A", "B"], "after": "owner"},
        ],
    }
    monkeypatch.setattr(hw, "load_config", lambda: cfg)
    return cfg


def test_crud_with_extra_fields(client, admin_headers, hw_config):
    r = client.post("/api/hardware-issues", headers=admin_headers, json={
        "summary": "电机异响", "extra_fields": {"factory": "工厂A", "grade": "A"}})
    assert r.status_code == 200, r.text
    row = r.json()
    assert row["extra_fields"] == {"factory": "工厂A", "grade": "A"}

    r = client.put(f"/api/hardware-issues/{row['id']}", headers=admin_headers, json={
        "version": row["version"], "extra_fields": {"factory": "工厂B"}})
    assert r.status_code == 200 and r.json()["extra_fields"] == {"factory": "工厂B"}

    # 乐观锁：旧 version 再改 → 409
    r = client.put(f"/api/hardware-issues/{row['id']}", headers=admin_headers, json={
        "version": row["version"], "summary": "x"})
    assert r.status_code == 409


def test_machine_summary_excludes_na(client, admin_headers, hw_config, machine_id):
    base = client.get("/api/hardware-issues/machine-summary", headers=admin_headers).json()
    base_s = base.get(str(machine_id), {"total": 0, "done": 0})
    for cell in ("已清零", "未清零", "不涉及"):
        r = client.post("/api/hardware-issues", headers=admin_headers, json={
            "summary": f"na测试-{cell}", "machine_cells": {str(machine_id): cell}})
        assert r.status_code == 200, r.text
    after = client.get("/api/hardware-issues/machine-summary", headers=admin_headers).json()
    s = after[str(machine_id)]
    assert s["total"] - base_s["total"] == 2   # 不涉及不进分母
    assert s["done"] - base_s["done"] == 1


def test_export_orders_custom_columns(client, admin_headers, hw_config):
    exp = client.get("/api/hardware-issues/export.xlsx", headers=admin_headers)
    assert exp.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(exp.content)).active
    hdr = [c.value for c in ws[1]]
    assert hdr.index("负责工厂") < hdr.index("来源")          # __start__ 在最前
    assert hdr.index("等级") == hdr.index("责任人") + 1       # after=owner 紧随责任人


def test_import_roundtrip(client, admin_headers, hw_config, machine_id):
    machine_no = client.get("/api/customer-status", headers=admin_headers).json()[0]["machine_id"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["编号", "来源", "问题单号", "问题简述", "负责工厂", "等级", machine_no])
    ws.append([1, "设计缺陷", "HW-T1", "导入往返", "工厂C", "B", "已清零"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/api/hardware-issues/import", headers=admin_headers,
                    files={"file": ("t.xlsx", buf, "application/octet-stream")})
    res = r.json()
    assert res["created"] == 1 and not res["errors"], res
    rows = [x for x in client.get("/api/hardware-issues", headers=admin_headers).json()
            if x["issue_ref"] == "HW-T1"]
    assert rows and rows[0]["extra_fields"] == {"factory": "工厂C", "grade": "B"}
    assert rows[0]["machine_cells"] == {str(machine_id): "已清零"}
