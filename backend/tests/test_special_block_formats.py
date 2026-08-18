"""自定义分段的两项扩展：里程碑分段，以及单元格 / 文本框的字体格式。

盯的是同一类回归——**页面能设置、导出却看不见**。分段形态与单元格格式各有三个
出口（详情页、周报 HTML、Excel），少接一处不会报错，只会让人在汇报当天发现
"页面上是宋体加粗红字，发出去的周报是默认字体"。所以这里对每种能力都同时验
周报文本、周报 HTML 与 Excel 三条路径。
"""
import io
import json

import pytest
from openpyxl import load_workbook


@pytest.fixture()
def sid(client, admin_headers):
    """一个干净的空专项（不套模板，分段由测试自己挂）。"""
    return client.post("/api/specials", headers=admin_headers, json={
        "name": "分段格式验证", "kind": "special",
    }).json()["id"]


def _content(client, headers, sid):
    return client.get(f"/api/specials/{sid}", headers=headers).json()["content"]


def _put_blocks(client, headers, sid, blocks):
    ver = _content(client, headers, sid)["version"]
    resp = client.put(f"/api/specials/{sid}/content", headers=headers, json={
        "version": ver,
        "extra_grids_json": json.dumps(blocks, ensure_ascii=False),
    })
    assert resp.status_code == 200, resp.text
    return resp.json()


def _draft(client, headers, sid):
    return client.get(f"/api/specials/{sid}/report-draft", headers=headers).json()["body"]


def _eml(client, headers, sid):
    resp = client.post(f"/api/specials/{sid}/report.eml", headers=headers, json={})
    assert resp.status_code == 200
    return resp.content.decode("utf-8", "replace")


def _xlsx(client, headers, sid):
    resp = client.get(f"/api/specials/{sid}/export.xlsx", headers=headers)
    assert resp.status_code == 200
    return load_workbook(io.BytesIO(resp.content))


# ─── 里程碑分段 ────────────────────────────────────────────────────

def test_milestone_block_reaches_report_and_export(client, admin_headers, sid):
    _put_blocks(client, admin_headers, sid, [{
        "gid": "msblock01", "kind": "milestones", "title": "专项测试计划",
        "milestones": [
            {"name": "冒烟测试完成", "date": "2026-09-01", "status": "done"},
            {"name": "回归测试完成", "date": "2026-09-20", "status": "in_progress"},
        ],
    }])

    body = _draft(client, admin_headers, sid)
    assert "专项测试计划" in body
    assert "冒烟测试完成" in body and "2026-09-01" in body
    # 状态用中文档位而不是 done/in_progress——周报是给人看的
    assert "[已完成]" in body and "[进行中]" in body

    eml = _eml(client, admin_headers, sid)
    assert "专项测试计划" in eml and "回归测试完成" in eml

    heads = [str(c[0].value) for c in _xlsx(client, admin_headers, sid)["专项"]
             .iter_rows(max_col=1) if c[0].value]
    assert any("专项测试计划" in h for h in heads), heads


def test_empty_milestone_block_takes_no_section_number(client, admin_headers, sid):
    """空分段整段跳过——这条口径页面、周报、Excel 必须一致。"""
    _put_blocks(client, admin_headers, sid, [
        {"gid": "msempty01", "kind": "milestones", "title": "还没排的计划",
         "milestones": []},
        {"gid": "txtblock1", "kind": "text", "title": "总体测试策略",
         "html": "<p>分三轮：冒烟 → 功能 → 回归</p>"},
    ])
    ver = _content(client, admin_headers, sid)["version"]
    client.put(f"/api/specials/{sid}/content", headers=admin_headers,
               json={"version": ver, "goal": "<p>目标正文</p>"})

    body = _draft(client, admin_headers, sid)
    assert "还没排的计划" not in body
    # 空段不占编号：目标是一、测试策略就该是二
    assert "一、专项目标" in body and "二、总体测试策略" in body

    heads = [str(c[0].value) for c in _xlsx(client, admin_headers, sid)["专项"]
             .iter_rows(max_col=1) if c[0].value]
    assert "二、总体测试策略" in heads, heads
    assert not any("还没排的计划" in h for h in heads)


# ─── 单元格格式 ────────────────────────────────────────────────────

def _fmt_grid():
    def cell(text, **fmt):
        return {"text": text, "align": "left", "color": "", "bg": "", **fmt}
    return {
        "gid": "fmtgrid01", "kind": "grid", "title": "解决方案验收标准",
        "headers": [{"text": "验收项", "colspan": 1, "align": "center", "font": "simsun"},
                    {"text": "达成情况", "colspan": 1, "align": "center"}],
        "colTypes": ["text", "light"],
        "colOptions": [[], ["绿", "黄", "红"]],
        "colWidths": [200, 90],
        "rows": [[
            cell("端到端联调", font="simsun", size=18, bold=True, italic=True,
                 underline=True, color="#C7000B", bg="#FFF7E6"),
            cell("绿"),
        ]],
    }


def test_cell_format_reaches_report_html(client, admin_headers, sid):
    _put_blocks(client, admin_headers, sid, [_fmt_grid()])
    eml = _eml(client, admin_headers, sid)

    assert "端到端联调" in eml
    for css in ("font-weight:700", "font-style:italic", "text-decoration:underline",
                "font-size:18px", "color:#C7000B", "background:#FFF7E6"):
        assert css in eml, f"{css} 没进周报 HTML"
    assert "SimSun" in eml, "字体没进周报 HTML"
    # 点灯列的着色仍然压过单元格自己的字色
    assert "F0F9EB" in eml


def test_cell_format_reaches_xlsx(client, admin_headers, sid):
    _put_blocks(client, admin_headers, sid, [_fmt_grid()])
    wb = _xlsx(client, admin_headers, sid)
    # 自定义表格走独立工作表
    sheet = next(n for n in wb.sheetnames if "验收标准" in n)
    ws = wb[sheet]
    cell = next(c for row in ws.iter_rows() for c in row if c.value == "端到端联调")

    assert cell.font.name == "宋体"
    assert cell.font.size == pytest.approx(13.5)      # 18px × 0.75
    assert cell.font.bold and cell.font.italic
    assert cell.font.underline == "single"
    assert cell.font.color.rgb.endswith("C7000B")
    assert cell.fill.fgColor.rgb.endswith("FFF7E6")


def test_illegal_font_and_size_are_dropped_not_rendered(client, admin_headers, sid):
    """白名单外的字体/字号一律退回默认，而不是原样拼进 style 或塞给 openpyxl。"""
    grid = _fmt_grid()
    grid["rows"][0][0].update({"font": "comic-sans", "size": 999,
                               "color": "red; background: url(x)"})
    _put_blocks(client, admin_headers, sid, [grid])

    eml = _eml(client, admin_headers, sid)
    assert "comic-sans" not in eml
    assert "font-size:999" not in eml
    assert "url(x)" not in eml, "字色没做校验就拼进了 style"

    wb = _xlsx(client, admin_headers, sid)
    ws = wb[next(n for n in wb.sheetnames if "验收标准" in n)]
    cell = next(c for row in ws.iter_rows() for c in row if c.value == "端到端联调")
    assert cell.font.name == "微软雅黑" and cell.font.size == 10


# ─── 文本框的 HTML 清洗 ────────────────────────────────────────────

def test_text_block_html_is_sanitized_on_save(client, admin_headers, sid):
    """文本框的 HTML 页面上是 v-html 渲染的，必须在服务端落库前清洗。"""
    saved = _put_blocks(client, admin_headers, sid, [{
        "gid": "xssblock1", "kind": "text", "title": "总体测试策略",
        "html": ('<p style="text-align:center">分三轮'
                 '<span style="font-family:SimSun, 宋体, serif">冒烟</span>'
                 '</p><script>alert(1)</script>'
                 '<img src=x onerror="alert(2)">'),
    }])
    html = json.loads(saved["extra_grids_json"])[0]["html"]

    assert "<script" not in html and "alert(1)" not in html
    assert "onerror" not in html and "<img" not in html
    # 合法的排版与字体要留着，否则清洗就成了"格式功能不能用"
    assert "text-align: center" in html
    assert "SimSun" in html


def test_text_block_keeps_list_markup(client, admin_headers, sid):
    """列表按钮产出的 ul/li 要能过清洗——否则周报里项目符号被静默抹平。"""
    saved = _put_blocks(client, admin_headers, sid, [{
        "gid": "listblk01", "kind": "text", "title": "总体测试策略",
        "html": "<ul><li>冒烟</li><li>回归</li></ul>",
    }])
    html = json.loads(saved["extra_grids_json"])[0]["html"]
    assert "<ul>" in html and html.count("<li>") == 2


# ─── 存量库的模板升级（0009）────────────────────────────────────────

def _load_migration():
    """按路径加载迁移模块——版本文件名不是合法的 import 路径。"""
    import importlib.util
    import pathlib
    path = (pathlib.Path(__file__).resolve().parent.parent
            / "alembic" / "versions" / "0009_solution_template_test_sections.py")
    spec = importlib.util.spec_from_file_location("mig0009", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def test_migration_patches_old_layout_in_place():
    """存量库里的老版式补三段：插在「测试点灯」之前，既有分段一个不动。"""
    mig = _load_migration()
    old = {
        "order": ["goal", "progress", "tpl:test-lights", "tpl:battlefield-risk",
                  "risks", "tpl:req-handover", "help"],
        "config": {"goal": {"title": "解决方案专项目标", "enabled": True}},
        "blocks": [
            {"tkey": "test-lights", "kind": "grid", "title": "测试详细进展和点灯",
             "headers": [{"text": "改过的列名", "colspan": 1, "align": "center"}]},
            {"tkey": "battlefield-risk", "kind": "grid", "title": "战场关键风险评估"},
            {"tkey": "req-handover", "kind": "grid", "title": "需求转测和规范度评估"},
        ],
    }
    new = mig._patched(old)

    assert new["order"] == [
        "goal", "progress",
        "tpl:acceptance", "tpl:test-strategy", "tpl:test-plan",
        "tpl:test-lights", "tpl:battlefield-risk",
        "risks", "tpl:req-handover", "help"]
    # admin 改过的列名不能被模板"修正"回去
    lights = next(b for b in new["blocks"] if b["tkey"] == "test-lights")
    assert lights["headers"][0]["text"] == "改过的列名"
    assert new["config"] == old["config"]

    # 幂等：已经补过的库再跑一次不重复插
    assert mig._patched(new) is None


def test_migration_skips_templates_that_already_have_the_sections():
    mig = _load_migration()
    layout = {"order": ["tpl:acceptance"], "config": {},
              "blocks": [{"tkey": t} for t in
                         ("acceptance", "test-strategy", "test-plan", "test-lights")]}
    assert mig._patched(layout) is None


def test_migration_appends_when_anchor_section_was_removed():
    """admin 删掉了「测试点灯」分段：没有锚点就补在末尾，而不是丢掉。"""
    mig = _load_migration()
    new = mig._patched({"order": ["goal", "help"], "config": {},
                        "blocks": [{"tkey": "battlefield-risk"}]})
    assert new["order"] == ["goal", "help", "tpl:acceptance",
                            "tpl:test-strategy", "tpl:test-plan"]


def test_builtin_rich_fields_are_sanitized_on_save(client, admin_headers, sid):
    """目标 / 整体进展 / 求助 也是 v-html 渲染的，同样要在入口清洗。"""
    ver = _content(client, admin_headers, sid)["version"]
    resp = client.put(f"/api/specials/{sid}/content", headers=admin_headers, json={
        "version": ver,
        "goal": '<b style="font-family:KaiTi, 楷体, serif">目标</b><script>alert(1)</script>',
        "help_request": '<div onclick="steal()">求助</div>',
    })
    assert resp.status_code == 200, resp.text
    got = resp.json()
    assert "<script" not in got["goal"] and "alert(1)" not in got["goal"]
    assert "KaiTi" in got["goal"], "合法字体被误伤"
    assert "onclick" not in got["help_request"] and "求助" in got["help_request"]


def test_item_rich_fields_are_sanitized_on_save(client, admin_headers, sid):
    """事务 / 风险的富文本同理——它们也进详情页的 v-html。"""
    resp = client.post(f"/api/specials/{sid}/tasks", headers=admin_headers, json={
        "content": '<span style="color:#C7000B">联调</span><script>alert(1)</script>',
        "progress": '<img src=x onerror="alert(2)">进展',
    })
    assert resp.status_code == 200, resp.text
    task = resp.json()
    assert "alert(1)" not in task["content"] and "#C7000B" in task["content"]
    assert "onerror" not in task["progress"] and "进展" in task["progress"]

    resp = client.put(f"/api/specials/tasks/{task['id']}", headers=admin_headers,
                      json={"content": "<u>改过</u><script>alert(3)</script>"})
    assert resp.status_code == 200, resp.text
    assert "alert(3)" not in resp.json()["content"]
    assert "<u>改过</u>" in resp.json()["content"]
