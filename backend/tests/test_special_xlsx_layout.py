"""专项 Excel 导出的**版式**回归。

版式跑偏和配色跑偏一样，是"看不出报错"的那类问题：文件照样生成、照样打得开，
只是里程碑上的日期串到隔壁去了、英文被从单词中间劈开、一份周报打开是四五个页签。
没人会把这些当 bug 报上来，所以把口径写成断言。

配色那一头见 [test_xlsx_style.py](test_xlsx_style.py)。
"""
import json

import pytest
import special_layout
import xlsx_utils


class _Any:
    """缺什么属性给什么（同 test_xlsx_style，build_special_xlsx 读的字段很多）。"""

    def __init__(self, **kw):
        self.__dict__.update(kw)

    def __getattr__(self, _):
        return ""


def _grid_block(gid="g1", title="分模块进展"):
    return {
        "gid": gid, "kind": "grid", "title": title,
        "headers": [{"text": t, "colspan": 1} for t in ("模块", "负责人", "风险", "备注")],
        "colTypes": ["text", "text", "light", "text"],
        "colWidths": [140, 90, 70, 300],
        "rows": [[{"text": "采集"}, {"text": "张三"}, {"text": "绿"}, {"text": "接口已联调"}],
                 [{"text": "算法"}, {"text": "王五"}, {"text": "红"}, {"text": "精度未达标"}]],
    }


def _special(*, milestones=None, blocks=(), order=None):
    content = _Any(goal="目标", progress_summary="进展",
                   milestones_json=json.dumps(milestones or []),
                   section_order_json=json.dumps(order or ["goal", "plan", "tasks"]),
                   section_config_json="{}",
                   extra_grids_json=json.dumps(list(blocks)))
    task = _Any(content="事务一", progress="推进中", owner="张三",
                status="open", planned_close_date="2026-09-20", sort_order=1, id=1)
    return _Any(id=1, kind="special", name="某专项", owner="李波",
                content=content, tasks=[task], risks=[])


def _sheet(sp):
    from openpyxl import load_workbook
    return load_workbook(xlsx_utils.build_special_xlsx(sp))


def _texts(ws):
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


def _font_or_skip(size=15, bold=False):
    font = xlsx_utils._load_pil_font(size, bold=bold)
    if font is None:
        pytest.skip("这台机器上没有可渲染中文的字体（导出会自动退回表格形式）")
    return font


# ─── 一个专项 = 一张表 ─────────────────────────────────────────────────────
def test_one_special_exports_one_sheet():
    """自定义表格**内联在对应章节下**，不再另开页签。

    以前它因为列宽对不上被丢到独立工作表、主表只留一行「→ 见工作表「X」」：
    一份周报打开是四五个页签，顺序、章节号和上下文全断在那一行指引上。
    """
    wb = _sheet(_special(blocks=[_grid_block()],
                         order=["goal", "grid:g1", "tasks"]))
    assert wb.sheetnames == ["专项"], wb.sheetnames
    txt = _texts(wb.active)
    assert "二、分模块进展" in txt, "自定义分段要占一个章节号"
    assert "接口已联调" in txt, "自定义表格的内容应该就在主表里"


def test_wide_grid_still_fits_the_same_grid():
    """8 列的自由表格和 6 列的事务表并存在同一张表上，谁都不被压变形。"""
    wide = _grid_block(gid="g2", title="八列宽表")
    wide["headers"] = [{"text": f"列{i}", "colspan": 1} for i in range(1, 9)]
    wide["colTypes"] = ["text"] * 8
    wide["colWidths"] = [60, 200, 80, 90, 90, 70, 60, 260]
    wide["rows"] = [[{"text": f"v{i}"} for i in range(8)]]
    ws = _sheet(_special(blocks=[wide], order=["grid:g2", "tasks"])).active
    # 每张表都铺满同一张 36 列网格（值只落在合并区左上角，所以看合并区右边界）
    assert max(rng.max_col for rng in ws.merged_cells.ranges) == xlsx_utils._NCOL
    assert "v7" in _texts(ws)


def test_column_ratios_fill_the_grid_exactly():
    """列宽传**比例**、归一化到正好铺满，不留半列也不越界。"""
    for ratios in ([6, 36, 30, 12, 14, 10], [1, 1, 1], [140, 90, 70, 300], [1] * 12):
        spans = xlsx_utils._spans_from_ratios(ratios)
        assert len(spans) == len(ratios)
        assert spans[0][0] == 1 and spans[-1][1] == xlsx_utils._NCOL
        for (a, b), (c, _d) in zip(spans, spans[1:]):
            assert a <= b and b + 1 == c, "相邻列必须首尾相接"


# ─── 里程碑：日期与英文 ────────────────────────────────────────────────────
@pytest.mark.parametrize("raw,want", [
    ("2026-05-20T00:00:00", "2026-05-20"),   # 带时分秒
    ("2026/7/1", "2026-07-01"),              # 斜杠 + 一位月份
    ("2026年9月15日", "2026-09-15"),
    ("1767225600000", "2026-01-01"),         # 毫秒时间戳
    ("2026-01-15", "2026-01-15"),
    ("6月中旬", "6月中旬"),                    # 认不出来的原样留着，别丢
    ("", ""),
])
def test_milestone_date_is_normalized(raw, want):
    assert special_layout.milestone_date_text(raw) == want


def test_report_and_excel_read_the_date_the_same_way(client, admin_headers):
    """周报与 Excel 走同一份归一。

    各写一份的表现是同一个里程碑在周报里是 "2026-05-20T00:00:00"、
    在 Excel 里是 "2026-05-20"，而两边看着都对。
    """
    from routers import specials
    ms = [{"name": "转测试", "date": "2026-05-20T00:00:00", "status": "done"}]
    assert "2026-05-20" in specials._milestone_table(ms)
    assert "00:00:00" not in specials._milestone_table(ms)
    assert "2026-05-20" in "".join(specials._milestone_lines(ms))


def test_milestone_table_fallback_also_normalizes(monkeypatch):
    """没有中文字体时退回表格形式，日期同样要归一。"""
    monkeypatch.setattr(xlsx_utils, "_render_milestone_image", lambda _ms: None)
    ws = _sheet(_special(milestones=[{"name": "转测试", "date": "2026/7/1",
                                      "status": "done"}])).active
    assert "2026-07-01" in _texts(ws)


def test_english_words_are_not_split_mid_word():
    """逐字符折行会把 "Release" 断成「Releas / e」——中文看不出问题，英文一眼就是坏的。"""
    from PIL import Image, ImageDraw
    font = _font_or_skip()
    draw = ImageDraw.Draw(Image.new("RGB", (8, 8)))
    lines = xlsx_utils._wrap_by_width(draw, "Alpha 版本发布 Release", font, 120)
    assert len(lines) > 1, "这个宽度下本来就该折行，否则这条断言没意义"
    for word in ("Alpha", "Release"):
        assert any(word in ln for ln in lines), f"{word} 被从中间劈开了：{lines}"


def test_a_single_overlong_word_is_still_broken():
    """一个词本身就装不下一行时才允许劈开——否则它会横着顶出图外。"""
    from PIL import Image, ImageDraw
    font = _font_or_skip()
    draw = ImageDraw.Draw(Image.new("RGB", (8, 8)))
    lines = xlsx_utils._wrap_by_width(draw, "Supercalifragilisticexpialidocious", font, 60)
    assert len(lines) > 1
    for ln in lines:
        assert draw.textbbox((0, 0), ln, font=font)[2] <= 60


# ─── 里程碑：版面 ──────────────────────────────────────────────────────────
def test_timeline_wraps_instead_of_growing_wider():
    """里程碑多了要**换行摆**，不是把图一路加宽。

    Excel 里图片不跟着列走：画到 3000px 宽就是横着甩出表格外面老远，
    而文件能生成、能打开，只有把表拉到最右边才看得见。
    """
    _font_or_skip()
    ms = [{"name": f"里程碑{i}", "date": "2026-01-%02d" % (i % 28 + 1), "status": "done"}
          for i in range(20)]
    img = xlsx_utils._render_milestone_image(ms)
    assert img is not None
    assert img.width <= xlsx_utils._SHEET_PX, f"轴太宽了：{img.width}"
    assert img.height > 200, "换行之后应该是往下长"


def test_slot_widens_for_long_dates():
    """槽宽由**最长的那一格**算出来，不是写死的：写死的话长日期直接压到隔壁节点上。"""
    _font_or_skip()
    short = xlsx_utils._render_milestone_image(
        [{"name": "A", "date": "2026-01-01", "status": "done"}] * 3)
    long = xlsx_utils._render_milestone_image(
        [{"name": "A", "date": "2026年第一季度末（待定）", "status": "done"}] * 3)
    assert long.width > short.width


def test_image_reserves_rows_matching_its_height():
    """图片是浮在格子上的，预留行数要和它的高度对得上。

    多留会在图后面空出半屏白（看着像"这一段没导出来"），少留会让下一段的标题
    被图压住。
    """
    _font_or_skip()
    ms = [{"name": f"里程碑{i}", "date": "2026-01-01", "status": "done"} for i in range(4)]
    sp = _special(milestones=ms, order=["plan", "tasks"])
    ws = _sheet(sp).active
    img = ws._images[0]
    reserved = xlsx_utils._rows_for_px(img.height)
    assert reserved * 20 >= img.height, "预留的行装不下这张图，下一段会被压住"
    assert reserved * 20 <= img.height + 60, "预留过头，图后面会空出一片白"


# ─── 边框 ─────────────────────────────────────────────────────────────────
# 这一节**必须读 xlsx 的原始 XML**，不能拿 openpyxl 读回来的对象验证：
# 它的 reader 会自己把合并区左上角的边框铺到四周，于是读回来看着是完整的框、
# 而文件里只有左上角那一格——Excel 打开就是「每段后面一条小短线、右边没有框」。
def _raw_borders(sp):
    """{行号: {列字母: {"left"/"right"/"top"/"bottom": 线型或 None}}}，直接解自文件。"""
    import re
    import zipfile

    z = zipfile.ZipFile(xlsx_utils.build_special_xlsx(sp))
    sheet = z.read("xl/worksheets/sheet1.xml").decode()
    styles = z.read("xl/styles.xml").decode()
    xfs = re.search(r'<cellXfs count="\d+">(.*?)</cellXfs>', styles, re.S).group(1)
    xf_border = [int(m or 0) for m in re.findall(r'<xf [^>]*borderId="(\d+)"', xfs)]
    blocks = re.findall(
        r"<border[ >].*?</border>|<border/>",
        re.search(r'<borders count="\d+">(.*?)</borders>', styles, re.S).group(1), re.S)

    def sides(idx):
        b = blocks[idx]
        out = {}
        for side in ("left", "right", "top", "bottom"):
            m = re.search(r'<%s style="(\w+)"' % side, b)
            out[side] = m.group(1) if m else None
        return out

    rows = {}
    for row_xml in re.finditer(r'<row r="(\d+)"[^>]*>(.*?)</row>', sheet, re.S):
        rows[int(row_xml.group(1))] = {
            col: sides(xf_border[int(sid or 0)])
            for col, sid in re.findall(r'<c r="([A-Z]+)\d+"(?: s="(\d+)")?',
                                       row_xml.group(2))}
    return rows


def test_merged_rows_are_written_cell_by_cell():
    """整幅合并的行（章节标题、正文段落）在文件里必须有 36 个格子。

    只写左上角那一格的话，Excel 画出来的框只在 A 列那么宽处露出一小截
    ——就是"每个段落后面一条小短线"，而外框整个是没有的。
    """
    rows = _raw_borders(_special())
    written = {r: c for r, c in rows.items() if c}
    assert written, "一行都没写？"
    for rown, cells in written.items():
        assert len(cells) == xlsx_utils._NCOL, f"第 {rown} 行只写了 {len(cells)} 格"


def test_every_section_has_a_complete_outer_frame():
    """每个分段（标题行 + 内容）压一个完整外框，四条边都要真的落在文件里。"""
    rows = _raw_borders(_special())
    last = "A" * 0 or _col_letter(xlsx_utils._NCOL)
    tops = [r for r, c in sorted(rows.items())
            if c and all(c[col]["top"] == "medium" for col in ("A", "R", last))]
    assert tops, "分段顶边没铺满整幅"
    top = tops[0]
    assert rows[top]["A"]["left"] == "medium", "左边没有"
    assert rows[top][last]["right"] == "medium", "**右边没有**——只写左上角时就是这个症状"
    bottoms = [r for r in sorted(rows) if r > top and rows[r]
               and all(v["bottom"] == "medium" for v in rows[r].values())]
    assert bottoms, "分段底边没铺满整幅"
    bot = bottoms[0]
    for r in range(top, bot + 1):
        assert rows[r]["A"]["left"] == "medium" and rows[r][last]["right"] == "medium", \
            f"第 {r} 行的左右边断了"


def _col_letter(n):
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)
