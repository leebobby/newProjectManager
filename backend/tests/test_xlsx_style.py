"""Excel 导出配色回归。

配色跑偏是**看不出报错**的那类问题：文件照样生成、照样能打开，只是
"这份表看着不像我们的材料"，没人会当 bug 报上来。所以把口径写成断言。

取色逻辑见 [brand.py](../brand.py)：红定位、蓝分层、饱和色点灯，三者各管一件事。
"""
import brand
import pptx_utils as PU
import xlsx_io
import xlsx_utils
from openpyxl import Workbook


def _rgb(color):
    """openpyxl 读回来的颜色是带 alpha 的 8 位串（00RRGGBB）。"""
    val = getattr(color, "rgb", None)
    return val[-6:] if isinstance(val, str) else None


# ─── 单一来源 ───────────────────────────────────────────────────────────────
def test_ppt_and_excel_share_one_palette():
    """PPT 与 Excel 必须是同一套色。

    历史上清单类 Excel 走品牌蓝 #4073BA、专项 Excel 走华为红、PPT 又是另一套红，
    同一批数据导出的三个文件是三个配色——而每一份单独看都挺正常。
    这条断言就是拦"又有人在某个模块里另写了一个颜色"。
    """
    assert str(PU._HEADER_BG) == brand.HEADER_BG
    assert str(PU._BRAND) == brand.BRAND
    assert str(PU._ZEBRA) == brand.ZEBRA
    assert xlsx_io.BRAND == brand.HEADER_BG
    assert xlsx_utils._HEADER_BG == brand.HEADER_BG
    assert xlsx_utils._BRAND == brand.BRAND
    assert {k: str(v) for k, v in PU._STATUS_FILLS.items()} == brand.STATUS_FILLS
    # 点灯（自由表格的点灯列 / 专项总览的风险灯）也只有一份：Excel 与 PPT
    # 各写一份字面量的表现是同一盏灯在两份文件里颜色不一样，而两份单独看都正常
    assert xlsx_utils._LIGHT_FILL == brand.LIGHT_FILLS
    assert xlsx_utils._LIGHT_FONT == brand.LIGHT_TEXTS
    assert {k: str(v) for k, v in PU._LIGHT_FILLS.items()} == brand.LIGHT_FILLS
    assert {k: str(v) for k, v in PU._LIGHT_TEXTS.items()} == brand.LIGHT_TEXTS
    # 但这两套**不能合并**：一套认进展状态词，一套认用户拨的灯
    assert set(brand.LIGHT_FILLS) & set(brand.STATUS_FILLS) == set()


def test_status_match_is_exact_not_substring():
    """"已完成三个模块联调"整格染绿是错的——它其实还在进行中。"""
    assert brand.status_style("已完成")[0] == "92D050"
    assert brand.status_style("  已完成  ")[0] == "92D050", "两边空格要容忍"
    assert brand.status_style("已完成三个模块联调")[0] is None
    assert brand.status_style("")[0] is None
    # 「未开始 / 不涉及」故意不点灯：它们说的是"这里没有进展"，
    # 上了底色会跟真有状态的格子一样抢眼
    fill, font, bold = brand.status_style("未开始")
    assert fill is None and font == brand.STATUS_TEXT["未开始"] and not bold


# ─── 清单类导出（xlsx_io） ──────────────────────────────────────────────────
def _sheet(rows, **kw):
    ws = Workbook().active
    xlsx_io.style_header(ws, ["名称", "进展", "备注"])
    for r in rows:
        ws.append(r)
    xlsx_io.beautify(ws, **kw)
    return ws


def test_list_export_header_is_light_blue_with_dark_text():
    """表头不再是红底/蓝底白字：一份文件里的红只留给标题。"""
    ws = _sheet([["A", "已完成", ""]])
    head = ws.cell(1, 1)
    assert _rgb(head.fill.fgColor) == brand.HEADER_BG
    assert _rgb(head.font.color) == brand.HEADER_TEXT
    assert head.font.bold


def test_list_export_lights_status_cells():
    """状态格上底色。只染字色的话，一列小字得逐格读才分得出来。"""
    ws = _sheet([["A", "已完成", "已完成三个模块联调"],
                 ["B", "已延期", ""],
                 ["C", "未开始", ""]])
    assert _rgb(ws.cell(2, 2).fill.fgColor) == "92D050"
    assert _rgb(ws.cell(3, 2).fill.fgColor) == "FF9999"
    # 备注列里那句话不是状态词，不能跟着染绿
    assert _rgb(ws.cell(2, 3).fill.fgColor) != "92D050"
    # 未开始不点灯，但字压灰
    assert _rgb(ws.cell(4, 2).fill.fgColor) != "92D050"
    assert _rgb(ws.cell(4, 2).font.color) == brand.STATUS_TEXT["未开始"]


def test_list_export_can_turn_status_light_off():
    """个别表里状态词只是普通文本，要有关掉的入口。"""
    ws = _sheet([["A", "已完成", ""]], light_status=False)
    assert _rgb(ws.cell(2, 2).fill.fgColor) != "92D050"


# ─── 专项周报（xlsx_utils） ─────────────────────────────────────────────────
class _Any:
    """缺什么属性给什么。build_special_xlsx 读的字段很多，
    逐个列出来的假对象每加一个字段就会碎一次。"""

    def __init__(self, **kw):
        self.__dict__.update(kw)

    def __getattr__(self, _):
        return ""


def _special_sheet():
    from openpyxl import load_workbook
    content = _Any(goal="目标内容", overall_progress="整体进展", milestones_json="[]")
    task = _Any(content="事务一", progress="推进中", owner="张三",
                status="进行中", sort_order=1, id=1)
    sp = _Any(kind="special", name="某专项", owner="李波",
              content=content, tasks=[task], risks=[])
    return load_workbook(xlsx_utils.build_special_xlsx(sp)).active


def test_special_report_title_is_red_text_not_a_red_banner():
    """标题红字 + 底下一条细红线，不是深红横幅压顶。

    横幅一铺，整份周报里最抢眼的是那两条红，而看的人要找的是表里的进展。
    """
    ws = _special_sheet()
    title = ws.cell(1, 1)
    assert "周报" in str(title.value)
    assert _rgb(title.font.color) == brand.BRAND
    assert _rgb(title.fill.fgColor) != brand.BRAND, "标题行不该整行铺红"

    # 整张表里被红填满的行只有那条线，而且必须细——粗了就退化回横幅
    red_rows = [r for r in range(1, 20)
                if _rgb(ws.cell(r, 1).fill.fgColor) == brand.BRAND]
    assert len(red_rows) == 1
    assert (ws.row_dimensions[red_rows[0]].height or 99) <= 4


def test_special_report_sections_and_headers_are_layered():
    """章节行中灰、表头浅蓝、正文白——三层各一个色，红不参与分层。"""
    ws = _special_sheet()
    fills = {str(ws.cell(r, 1).value): _rgb(ws.cell(r, 1).fill.fgColor)
             for r in range(1, 20) if ws.cell(r, 1).value}
    sections = [k for k in fills if k.startswith(("一、", "二、"))]
    assert sections, "应有章节行"
    for k in sections:
        assert fills[k] == brand.SECTION_BG, f"章节行 {k} 应为中灰"
    assert fills.get("序号") == brand.HEADER_BG, "表头应为浅蓝灰"


def test_special_report_lights_only_the_status_cell():
    """底色只上在状态那一格，不整行铺。

    整行绿看着像"这一行整体没问题"，而它表达的其实只是某一列填了状态。
    """
    ws = _special_sheet()
    row = next(r for r in range(1, 20) if str(ws.cell(r, 1).value) == "1")
    # 每个逻辑列在 36 列物理网格上占一段合并区，样式只落在左上角那一格
    lit = [c for c in range(1, xlsx_utils._NCOL + 1)
           if _rgb(ws.cell(row, c).fill.fgColor) in brand.STATUS_FILLS.values()]
    assert len(lit) == 1, "只有状态格该点灯"
    assert lit[0] == xlsx_utils._spans_from_ratios(xlsx_utils._SIX_RATIOS)[-1][0], \
        "点灯的应该是最后一列（状态）"
    assert str(ws.cell(row, lit[0]).value) == "进行中"
    assert _rgb(ws.cell(row, lit[0]).fill.fgColor) == "FFD966"
