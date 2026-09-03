"""Excel 导出工具：专项/攻关导出为美观、整洁、可直接使用的单页表格。

依赖：openpyxl

设计要点（解决"导出很乱"）：
- **一个专项导出成一张表**：所有分段（含自定义自由表格）按页面顺序往下排。
  以前自由表格被丢到独立工作表、主表只留一行"→ 见工作表「X」"，一份周报打开
  是四五个页签，顺序和上下文全断在那行指引上；
- 全表统一 **36 列等宽窄网格**，每张表按「列宽比例」合并到这张网格上——
  列宽一样宽之后，6 列的事务表与 5~8 列的自由表格才能并存而谁都不变形；
- 标题 / 章节 / 叙述段落横跨整幅合并，但样式**逐格写**：xlsx 文件里合并区只留
  左上角那一格，而 Excel 画合并区的边框是按边上每一格取的——只写左上角，
  那道线就只在 A 列那么宽处露出一小截，右边整个没有。见 `_frame()`；
- 单元格统一自动换行 + 按内容估算行高，长文本不再溢出/挤压；
- 打印按横向、缩放到一页宽，不然 36 列会被从中间劈成左右两叠。

配色取自 [brand.py](brand.py)，与 PPT / 清单类 Excel **同一套**：
红只用在报告主标题上（底下压一条细红线），章节行走中灰，表头走浅蓝灰，
状态格才上饱和色。原来是深红横幅压顶 + 红底白字表头 + 浅红斑马，
一份周报里红铺满了小半页，最抢眼的成了那些红条，而看的人要找的是表里的进展。
"""
import io
import math
import os
import pathlib
import re
from datetime import datetime
from typing import Optional

import brand
import enums
import special_layout
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 配色单一来源见 brand.py，**不要在这里另写颜色字面量**
_BRAND = brand.BRAND          # 华为红：只给报告主标题的文字和它底下那条线
_SECTION = brand.SECTION_BG   # 章节标题底色（中灰）
_HEADER_BG = brand.HEADER_BG  # 表头底色（浅蓝灰）
_ZEBRA = brand.ZEBRA          # 斑马底色
_BORDER_RGB = brand.BORDER

_FONT = "微软雅黑"

# ─── 物理列网格 ────────────────────────────────────────────────────────────
# **36 个等宽窄列**（每列 3 字符 ≈ 26px），所有表格按「列宽比例」合并到这张网格上。
# 原来是 6 个宽窄不一的列（[6,36,30,12,14,10]），于是自由表格（5~8 列、列宽各异）
# 塞不进来，只能丢到独立工作表里，一个专项导出成好几个页签。列宽一样宽之后，
# 6 列的事务表、3 列的里程碑表、8 列的自由表格可以并存在同一张表上而谁都不变形——
# 代价只是列标从 A~F 变成 A~AJ，而这份东西是给人看的报告，不是给人填的表。
# **列宽传比例、不传字符数**（同 pptx_utils 的 col_ratios）：加一列、改一个标题
# 就再也对不上，是历史上几张表宽度各不相同的原因。
_NCOL = 36
_BASE_W = 3.0
_SHEET_PX = 930          # 36 列大致占的像素宽，图片按它缩放/换行，别甩出表外

_MS_STATUS_LABEL = {
    "planning": "未开始", "in_progress": "进行中", "done": "已完成", "delayed": "已延期",
}
# 状态点灯：**给格子上底色**（brand.STATUS_FILLS），不是只染字色。
# 原来「已闭环/已完成」给整行铺浅绿、其余只染字色，两种表达混在一张表里：
# 绿行看着像"这一行整体没问题"，而它其实只是某一列填了已完成。
_STATUS_FONT = dict(brand.STATUS_TEXT)
_STATUS_FILL = dict(brand.STATUS_FILLS)

_thin = Side(style="thin", color=_BORDER_RGB)
_MEDIUM = Side(style="medium", color=_BORDER_RGB)
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# 事务/风险表的六列比例，与改造前的字符列宽 [6,36,30,12,14,10] 等价
_SIX_RATIOS = [6, 36, 30, 12, 14, 10]


def _spans_from_ratios(ratios, ncol: int = _NCOL):
    """列宽比例 → 每个逻辑列在物理网格里的 (起列, 止列)。

    归一化到正好铺满 ncol 列：差额补在**最宽的那列**上，比例走形最小。
    逻辑列比物理列还多时 1:1 摊开（此时表已经宽到没法再讲比例了）。
    """
    vals = [max(float(x or 0), 0.0) for x in ratios]
    n = len(vals)
    if n == 0:
        return []
    if n >= ncol:
        return [(i + 1, i + 1) for i in range(n)]
    total = sum(vals) or float(n)
    widths = [max(1, int(round(v / total * ncol))) for v in vals]
    while sum(widths) < ncol:
        widths[max(range(n), key=lambda k: widths[k])] += 1
    while sum(widths) > ncol:
        i = max(range(n), key=lambda k: widths[k])
        if widths[i] <= 1:
            break
        widths[i] -= 1
    spans, c = [], 1
    for w in widths:
        spans.append((c, c + w - 1))
        c += w
    return spans


def _with_sides(border, **sides):
    return Border(left=sides.get("left", border.left), right=sides.get("right", border.right),
                  top=sides.get("top", border.top), bottom=sides.get("bottom", border.bottom))


def _frame(ws, r1, r2, c1: int = 1, c2: int = _NCOL):
    """给 r1..r2 行、c1..c2 列这一块**外圈**压一道粗边。

    **必须逐格写在真正的边上**，不能只写合并区的左上角：xlsx 文件里合并区
    只留左上角那一格，而 Excel 画合并区的边框是**按边上每一格**取的——只写
    左上角，那道线就只在 A 列那么宽的地方露出一小截（"每段后面一条小短线"），
    右边则整个没有（"外框没有"）。

    **不要用 openpyxl 读回来的样子验证这件事**：它的 reader 会自己把左上角的
    边框铺到合并区四周，于是读回来看着是完整的框，而文件里根本没有。
    要验就解开 xlsx 数 `<c>` 标签，见 tests/test_special_xlsx_layout.py。
    """
    if r2 < r1 or c2 < c1:
        return
    for c in range(c1, c2 + 1):
        for r, side in ((r1, "top"), (r2, "bottom")):
            cell = ws.cell(row=r, column=c)
            cell.border = _with_sides(cell.border, **{side: _MEDIUM})
    for r in range(r1, r2 + 1):
        for c, side in ((c1, "left"), (c2, "right")):
            cell = ws.cell(row=r, column=c)
            cell.border = _with_sides(cell.border, **{side: _MEDIUM})


# 点灯列的底色/字色，与前端 RichGrid 及周报 HTML 的三档保持一致。
# 表在 brand.py（PPT 那边也要用同一套），这里不再另写一份字面量。
_LIGHT_FILL = dict(brand.LIGHT_FILLS)
_LIGHT_FONT = dict(brand.LIGHT_TEXTS)

# 分段图片 / 全景图的落盘根目录（与 routers/specials.py 的 UPLOAD_ROOT 同一处）
UPLOAD_ROOT = pathlib.Path(__file__).resolve().parent / "uploads" / "specials"


def _strip_html(s: str) -> str:
    if not s:
        return ""
    text = re.sub(r"<br\s*/?>", "\n", s, flags=re.I)
    text = re.sub(r"</\s*(p|div|h\d|li|tr)\s*>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = (text.replace("&nbsp;", " ").replace("&amp;", "&")
                .replace("&lt;", "<").replace("&gt;", ">").replace("&quot;", '"'))
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _kind_label(kind: str) -> str:
    return "攻关" if kind == "assault" else "专项"


def _disp_w(s: str) -> int:
    """显示宽度：CJK 记 2，其余记 1。"""
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in s)


def _cell_lines(text: str, capacity: int) -> int:
    cap = max(capacity, 4)
    lines = 0
    for seg in str(text or "").split("\n"):
        lines += max(1, math.ceil(_disp_w(seg) / cap))
    return max(1, lines)


def _rows_for_px(height_px: float) -> int:
    """图片占多少行。

    Excel 里图片是浮在格子上的、不占行——不预留就直接压住下一段的标题。
    默认行高 15 磅 ＝ **20px**（原来按 18px 算还另加 2 行，一张图后面能空出
    半屏白，看着像"这一段没导出来"）。多留一行做与下一段的间距。
    """
    return max(4, math.ceil(float(height_px or 0) / 20.0) + 1)


def _hex_to_rgb6(color: str, default: str = "262626") -> str:
    """'#C7000B' / 'C7000B' → 'C7000B'；空 / 非法 → default。"""
    s = (color or "").strip().lstrip("#")
    if len(s) == 6 and all(c in "0123456789abcdefABCDEF" for c in s):
        return s.upper()
    return default


# ─── 里程碑「图片」渲染（PIL）─────────────────────────────────────
# 里程碑导出为时间轴图片而非表格。字体在 Windows / Linux 上自动发现；
# 找不到能渲染中文的字体时整体放弃图片、退回表格形式（避免方块乱码）。
# 可用环境变量 APP_CJK_FONT 指定字体文件路径（部署机字体装在非常规目录时）。

_MS_DOT_RGB = {
    "planning": (192, 196, 204), "in_progress": (64, 158, 255),
    "done": (103, 194, 58), "delayed": (245, 108, 108),
}
_MS_LEGEND = [("planning", "未开始"), ("in_progress", "进行中"),
              ("done", "已完成"), ("delayed", "已延期")]

_FONT_CANDIDATES_REGULAR = [
    "C:/Windows/Fonts/msyh.ttc", "C:/Windows/Fonts/msyh.ttf",
    "C:/Windows/Fonts/simhei.ttf", "C:/Windows/Fonts/simsun.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf",
    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
    "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
    "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
    "/System/Library/Fonts/PingFang.ttc",
]
_FONT_CANDIDATES_BOLD = [
    "C:/Windows/Fonts/msyhbd.ttc", "C:/Windows/Fonts/simhei.ttf",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Bold.otf",
    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
]


# 常规目录下按文件名模式补充发现 CJK 字体（候选清单没命中时的兜底）
_FONT_SCAN_DIRS = [
    "C:/Windows/Fonts",
    "/usr/share/fonts", "/usr/local/share/fonts",
    "/System/Library/Fonts",
]
_FONT_SCAN_PATTERNS = (
    "msyh", "simhei", "simsun", "notosanscjk", "notoserifcjk",
    "sourcehansans", "wqy", "droidsansfallback", "pingfang", "harmonyos",
)


def _discover_cjk_font() -> Optional[str]:
    for d in _FONT_SCAN_DIRS:
        if not os.path.isdir(d):
            continue
        try:
            for root, _dirs, files in os.walk(d):
                for f in files:
                    low = f.lower()
                    if low.endswith((".ttf", ".ttc", ".otf")) and any(p in low for p in _FONT_SCAN_PATTERNS):
                        return os.path.join(root, f)
        except OSError:
            continue
    return None


def _can_render_cjk(font) -> bool:
    """字体是否真的带汉字字形（拿「中」探测；缺字形的字体 getmask 全空）。"""
    try:
        mask = font.getmask("中")
        return mask.getbbox() is not None
    except Exception:
        return False


def _load_pil_font(size: int, bold: bool = False):
    """按候选清单→环境变量→目录扫描找中文字体；全部落空返回 None（调用方退回表格），
    绝不回退 PIL 默认字体——它不含中文，画出来是方块。"""
    from PIL import ImageFont
    cands = ([os.environ.get("APP_CJK_FONT")] if os.environ.get("APP_CJK_FONT") else []) \
        + (_FONT_CANDIDATES_BOLD if bold else []) + _FONT_CANDIDATES_REGULAR
    for path in cands:
        if path and os.path.exists(path):
            try:
                font = ImageFont.truetype(path, size)
                if _can_render_cjk(font):
                    return font
            except OSError:
                continue
    found = _discover_cjk_font()
    if found:
        try:
            font = ImageFont.truetype(found, size)
            if _can_render_cjk(font):
                return font
        except OSError:
            pass
    return None


def _text_wh(draw, text: str, font):
    box = draw.textbbox((0, 0), text, font=font)
    return box[2] - box[0], box[3] - box[1]


_LATIN_RUN = re.compile(r"[0-9A-Za-z][0-9A-Za-z._/'+-]*")


def _break_tokens(s: str):
    """把一行切成「可断点」词元：西文/数字连成一串算一个词元，其余逐字。"""
    toks, i, n = [], 0, len(s)
    while i < n:
        m = _LATIN_RUN.match(s, i)
        if m:
            toks.append(m.group(0))
            i = m.end()
        else:
            toks.append(s[i])
            i += 1
    return toks


def _wrap_by_width(draw, text: str, font, max_w: int):
    """按显示宽度折行，**西文单词整体不拆**。

    原来是逐字符折行：中文看不出问题，英文一眼就是坏的——"Alpha 版本发布 Release"
    会断成「Releas / e」。只有单个词本身就装不下一行时才从中间劈开。
    """
    out = []
    for raw in str(text or "").split("\n"):
        cur = ""
        for tok in _break_tokens(raw):
            if cur and _text_wh(draw, cur + tok, font)[0] > max_w:
                out.append(cur)
                cur = tok if tok.strip() else ""
            else:
                cur += tok
            while _text_wh(draw, cur, font)[0] > max_w and len(cur) > 1:
                k = len(cur) - 1
                while k > 1 and _text_wh(draw, cur[:k], font)[0] > max_w:
                    k -= 1
                out.append(cur[:k])
                cur = cur[k:]
        out.append(cur)
    return [ln for ln in out if ln != ""] or [""]


def _ms_date_text(raw) -> str:
    """轴上的日期文字：归一交给 special_layout（周报也用同一份），空值写「未定」。"""
    return special_layout.milestone_date_text(raw) or "未定"


def _render_milestone_image(milestones):
    """把里程碑画成横向时间轴 PNG，返回 PIL.Image；PIL 不可用/出错时返回 None（调用方退回表格）。

    两条硬规则，都是"看着不像坏了"的那类问题：
    - **槽宽由最长的那一格算出来**，不是写死的 175px。写死的话长名字/长日期直接
      压到隔壁节点上，而图能生成、能打开，只是读不出哪个日期是谁的。
    - **超过表宽就换行摆**，不是把图一路加宽。Excel 里图片不跟着列走，
      画到 3000px 宽就是横着甩出表格外面老远，打印出来更是没边。
    """
    if not milestones:
        return None
    try:
        from PIL import Image, ImageDraw

        f_name = _load_pil_font(15, bold=True)
        f_date = _load_pil_font(12)
        f_legend = _load_pil_font(12)
        if not (f_name and f_date and f_legend):
            return None  # 没有可用中文字体 → 退回表格，避免方块乱码

        probe = ImageDraw.Draw(Image.new("RGB", (8, 8)))
        names = [str(m.get("name") or "") for m in milestones]
        dates = [_ms_date_text(m.get("date")) for m in milestones]
        n = len(milestones)

        # 槽宽：先让日期能一行放下（日期不该折行），再夹在 [150, 240] 之间
        date_w = max([_text_wh(probe, d, f_date)[0] for d in dates] + [70])
        slot = max(150, min(240, date_w + 26))
        name_w = slot - 14

        name_lines = [_wrap_by_width(probe, s, f_name, name_w) for s in names]
        date_lines = [_wrap_by_width(probe, s, f_date, slot - 8) for s in dates]
        lh_n = _text_wh(probe, "中", f_name)[1] + 4
        lh_d = _text_wh(probe, "中", f_date)[1] + 3
        name_h = max(len(x) for x in name_lines) * lh_n
        date_h = max(len(x) for x in date_lines) * lh_d

        side = 30
        cols = max(1, int((_SHEET_PX - side * 2) // slot))
        n_rows = math.ceil(n / cols)
        cols = math.ceil(n / n_rows)          # 均衡：6 个点不要摆成 5 + 1
        row_h = name_h + 20 + date_h + 26
        top = 12
        legend_h = 32
        width = side * 2 + cols * slot
        height = top + n_rows * row_h + legend_h

        img = Image.new("RGB", (width, height), "white")
        d = ImageDraw.Draw(img)

        def node_x(i):
            return side + slot * (i % cols) + slot / 2

        def baseline_y(i):
            return top + (i // cols) * row_h + name_h + 14

        # 轴线：每行一段，从该行第一个节点画到最后一个
        for r in range(n_rows):
            first, last = r * cols, min(n, (r + 1) * cols) - 1
            y = baseline_y(first)
            d.line([(node_x(first) - slot * 0.36, y), (node_x(last) + slot * 0.36, y)],
                   fill=(220, 223, 230), width=3)

        for i, m in enumerate(milestones):
            x, y = node_x(i), baseline_y(i)
            rgb = _MS_DOT_RGB.get(m.get("status", "planning"), _MS_DOT_RGB["planning"])
            # 名称（轴线上方，底对齐到轴线，加粗）
            ny = y - 16
            for ln in reversed(name_lines[i]):
                w, h = _text_wh(d, ln, f_name)
                d.text((x - w / 2, ny - h), ln, font=f_name, fill=(48, 49, 51))
                ny -= lh_n
            # 节点圆点（外圈白 + 彩色实心）
            r0 = 8
            d.ellipse([x - r0 - 2, y - r0 - 2, x + r0 + 2, y + r0 + 2], fill=(255, 255, 255))
            d.ellipse([x - r0, y - r0, x + r0, y + r0], fill=rgb)
            # 日期（轴线下方）
            dy = y + 14
            for ln in date_lines[i]:
                w, _h = _text_wh(d, ln, f_date)
                d.text((x - w / 2, dy), ln, font=f_date, fill=(144, 147, 153))
                dy += lh_d

        # 图例
        lx, ly = side, height - legend_h + 8
        for status, label in _MS_LEGEND:
            d.ellipse([lx, ly + 3, lx + 11, ly + 14], fill=_MS_DOT_RGB[status])
            d.text((lx + 16, ly), label, font=f_legend, fill=(96, 98, 102))
            lx += 16 + _text_wh(d, label, f_legend)[0] + 26

        return img
    except Exception:
        return None


def build_special_xlsx(special) -> io.BytesIO:
    """传入 Special ORM（含 content/tasks/risks），返回美观 xlsx 的 BytesIO。"""
    label = _kind_label(special.kind)
    content = special.content
    wb = Workbook()
    ws = wb.active
    ws.title = label
    ws.sheet_view.showGridLines = False

    for i in range(1, _NCOL + 1):
        ws.column_dimensions[get_column_letter(i)].width = _BASE_W
    # 打印版式：横向、按宽度缩放到一页宽。不设的话默认纵向 A4，一份 36 列的报告
    # 打出来被从中间劈成左右两叠，而在屏幕上看是好的。
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    # pending＝已经排到但还没落笔的章节标题，见 _flush_section
    state = {"row": 1, "pending": None}

    def _fill(r, c1, c2, color):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=color)

    def band(text, *, fill=None, font_color="262626", bold=False, size=11,
             align="left", height=20, italic=False, wrap=True):
        r = state["row"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_NCOL)
        ws.cell(row=r, column=1, value=text)
        font_ = Font(name=_FONT, bold=bold, size=size, color=font_color, italic=italic)
        align_ = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        # 整幅合并的行一律**逐格写**：文件里合并区只留左上角那一格，只写它的话
        # 后面再往这一行加边框就只能露出 A 列那么宽的一小截（见 _frame）
        for cc in range(1, _NCOL + 1):
            cell = ws.cell(row=r, column=cc)
            cell.font = font_
            cell.alignment = align_
        if fill:
            _fill(r, 1, _NCOL, fill)
        ws.row_dimensions[r].height = height
        state["row"] += 1

    def section(title):
        """把章节标题挂起，等真有内容落笔时再写（见 _flush_section）。"""
        state["pending"] = title

    def _flush_section():
        """写正文前补上待写的章节标题。

        「空分段不占章节编号」是页面 / 周报 / 导出三处共同的口径，可标题必须写在
        内容之前——所以先挂起、等第一笔内容落笔时再补。原先是先写标题、发现空了
        再补一行「—」，于是模板里多一个空分段，Excel 的章节号就和周报对不上，
        而这种错没人会当成 bug 去查。
        """
        title = state.get("pending")
        if not title:
            return
        state["pending"] = None
        band(title, fill=_SECTION, font_color=brand.HEADER_TEXT, bold=True, size=12, height=22)

    def narrative(text):
        _flush_section()
        r = state["row"]
        text = text or "—"
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_NCOL)
        c = ws.cell(row=r, column=1, value=text)
        font = Font(name=_FONT, size=11, color="262626")
        align = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        # **逐格写**：合并区在文件里只有左上角那一格，只给它上样式的话
        # Excel 那道框只在 A 列那么宽处露出一小截（见 _frame 的说明）
        for cc in range(1, _NCOL + 1):
            cell = ws.cell(row=r, column=cc)
            cell.font = font
            cell.alignment = align
            cell.border = _BORDER
        cap = int(_NCOL * _BASE_W) - 4
        ws.row_dimensions[r].height = max(20, min(260, _cell_lines(text, cap) * 16 + 4))
        state["row"] += 1

    def rule(color=_BRAND, height=3):
        """标题下的通栏细线。用一行极矮的填充行画，Excel 里没有真正的"横线"。"""
        r = state["row"]
        _fill(r, 1, _NCOL, color)
        ws.row_dimensions[r].height = height
        state["row"] += 1

    def gap():
        state["row"] += 1

    def table(col_specs, headers, rows, status_col=None, center_cols=None):
        """col_specs: [(c1,c2), ...] 每个逻辑列在 6 列网格中的物理列区间。
        status_col: 逻辑列下标（从 0 起），该列文字按状态着色 + 整行变浅绿。
        center_cols: 需要居中的逻辑列下标集合（默认 status / 第 0 列居中）。"""
        _flush_section()
        center = set(center_cols or [])
        # 列容量（字符单位）
        caps = [(c2 - c1 + 1) * _BASE_W for (c1, c2) in col_specs]

        # 表头
        r = state["row"]
        for j, (c1, c2) in enumerate(col_specs):
            if c2 > c1:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            for cc in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=cc, value=headers[j] if cc == c1 else None)
                cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
                cell.font = Font(name=_FONT, bold=True, color=brand.HEADER_TEXT, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = _BORDER
        ws.row_dimensions[r].height = 20
        state["row"] += 1

        # 数据行
        for i, dr in enumerate(rows):
            r = state["row"]
            status_txt = ""
            if status_col is not None and status_col < len(dr):
                status_txt = str(dr[status_col] or "").strip()
            zebra = (i % 2 == 1)
            # 底色只上在状态那一格上，不再整行铺色：整行绿看着像"这一行都没问题"，
            # 而它表达的其实只是某一列填了「已完成」
            row_fill = _ZEBRA if zebra else None
            max_lines = 1
            for j, (c1, c2) in enumerate(col_specs):
                if c2 > c1:
                    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
                val = "" if j >= len(dr) or dr[j] is None else str(dr[j])
                max_lines = max(max_lines, _cell_lines(val, caps[j]))
                is_status = (status_col is not None and j == status_col)
                lit, lit_font, lit_bold = brand.status_style(val) if is_status else (None, None, False)
                fcolor = lit_font or brand.TEXT
                halign = "center" if (j in center or is_status) else "left"
                cell_fill = lit or row_fill
                for cc in range(c1, c2 + 1):
                    cell = ws.cell(row=r, column=cc, value=val if cc == c1 else None)
                    cell.font = Font(name=_FONT, size=10, color=fcolor, bold=lit_bold)
                    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
                    cell.border = _BORDER
                    if cell_fill:
                        cell.fill = PatternFill("solid", fgColor=cell_fill)
            ws.row_dimensions[r].height = max(18, min(160, max_lines * 15 + 3))
            state["row"] += 1

    # ===== 标题区：红字标题 + 灰副标题 + 一条细红线 =====
    # 不再用深红横幅压顶。横幅一铺，整份周报里最抢眼的是那两条红，
    # 而看的人要找的是下面表里的进展。红只留给标题文字和这条线。
    band(f"【{label}周报】{special.name or ''}",
         font_color=_BRAND, bold=True, size=16, align="left", height=26)
    band(f"责任人：{special.owner or '-'}      导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
         font_color=brand.MUTED, size=10, align="left", height=16)
    rule()
    gap()

    kept_images = []  # 持有 BytesIO 引用直到 wb.save，避免被 GC
    # 事务/风险表：序号 / 内容 / 进展 / 责任人 / 闭环 / 状态
    six = _spans_from_ratios(_SIX_RATIOS)

    def render_milestones(milestones=None):
        """内置「计划」与自定义里程碑分段共用：两者只是取数不同，画法必须一样。"""
        if milestones is None:
            milestones = special_layout.loads(getattr(content, "milestones_json", None), [])
        milestones = [m for m in milestones if isinstance(m, dict)]
        if not milestones:
            return False
        _flush_section()
        ms_img = _render_milestone_image(milestones)
        if ms_img is not None:
            r = state["row"]
            bio = io.BytesIO()
            ms_img.save(bio, format="PNG")
            bio.seek(0)
            xi = XLImage(bio)
            if xi.width and xi.width > _SHEET_PX:      # 超宽就整体缩到表宽内
                ratio = _SHEET_PX / float(xi.width)
                xi.width = int(xi.width * ratio)
                xi.height = int(xi.height * ratio)
            ws.add_image(xi, f"A{r}")
            kept_images.append(bio)
            state["row"] += _rows_for_px(xi.height or ms_img.height)
        else:
            # PIL 不可用：退回表格形式
            rows = [[m.get("name", ""), _ms_date_text(m.get("date")),
                     _MS_STATUS_LABEL.get(m.get("status", "planning"), m.get("status", ""))]
                    for m in milestones]
            table(_spans_from_ratios([3, 1, 1]), ["里程碑", "日期", "状态"], rows,
                  status_col=2, center_cols={1, 2})
        return True

    def render_items(rows_src, content_header):
        if not rows_src:
            return False
        rows = []
        for idx, it in enumerate(rows_src, 1):
            st = "已闭环" if (it.status or "open") == "closed" else "进行中"
            rows.append([idx, _strip_html(it.content), _strip_html(it.progress),
                         it.owner or "", it.planned_close_date or "", st])
        table(six, ["序号", content_header, "当前进展", "责任人", "计划闭环", "状态"],
              rows, status_col=5, center_cols={0, 3, 4})
        return True

    def render_formation():
        obj = special_layout.loads(getattr(content, "formation_json", None), {})
        headers = [str(h) for h in (obj.get("headers") or [])]
        rows = [[_cell_str(c) for c in (r or [])] for r in (obj.get("rows") or [])]
        rows = [r for r in rows if any(x.strip() for x in r)]
        if not rows:
            return False
        ncol = max(len(headers), max(len(r) for r in rows))
        table(_spans_from_ratios([1] * ncol), headers or [""] * ncol, rows)
        return True

    def render_grid(block):
        """自定义表格（RichGrid）**内联进主表**，不再丢到独立工作表。

        列宽按 colWidths 的**比例**映射到 36 列物理网格上（同 pptx_utils 的 col_ratios），
        于是 5 列的表和 8 列的表可以并存在同一张表里而谁都不被压变形。
        单元格自带的字体/字号/字色/底色与点灯列原样带过来——那是用户在页面上一格
        一格调出来的，导出里丢掉等于"格式功能不能用"。
        """
        raw_headers = block.get("headers") or []
        raw_rows = block.get("rows") or []
        col_types = block.get("colTypes") or []
        if not _grid_cells(block)[1]:
            return False

        hdrs = []
        for h in raw_headers:
            if isinstance(h, dict):
                hdrs.append({**h, "text": str(h.get("text", "")),
                             "colspan": max(1, int(h.get("colspan") or 1)),
                             "align": h.get("align") or "center"})
            else:
                hdrs.append({"text": str(h), "colspan": 1, "align": "center"})
        body_cols = sum(h["colspan"] for h in hdrs)
        if body_cols <= 0:
            body_cols = max((len(r) for r in raw_rows if isinstance(r, list)), default=0)
            hdrs = [{"text": f"列{i + 1}", "colspan": 1, "align": "center"}
                    for i in range(body_cols)]
        if not body_cols:
            return False

        col_widths = block.get("colWidths") or []

        def _px(i, default=130.0):
            if i < len(col_widths):
                try:
                    return max(20.0, float(col_widths[i]))
                except (TypeError, ValueError):
                    return default
            return default

        spans = _spans_from_ratios([_px(i) for i in range(body_cols)])
        caps = [(c2 - c1 + 1) * _BASE_W for (c1, c2) in spans]

        _flush_section()
        r = state["row"]
        col = 0
        for h in hdrs:
            last = min(col + h["colspan"], body_cols) - 1
            c1, c2 = spans[col][0], spans[last][1]
            if c2 > c1:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            hf = _cell_font_spec(h)
            for cc in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=cc, value=h["text"] if cc == c1 else None)
                cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
                # 表头恒为浅蓝灰底黑字粗体：只跟随字体与字号，字色/底色不跟
                cell.font = Font(name=hf["name"], bold=True, color=brand.HEADER_TEXT,
                                 size=hf["size"], italic=hf["italic"],
                                 underline=hf["underline"])
                cell.alignment = Alignment(horizontal=h["align"], vertical="center",
                                           wrap_text=True)
                cell.border = _BORDER
            col = last + 1
        ws.row_dimensions[r].height = 20
        state["row"] += 1

        for i, row in enumerate(raw_rows):
            cells = row if isinstance(row, list) else []
            r = state["row"]
            zebra = (i % 2 == 1)
            max_lines = 1
            for j, (c1, c2) in enumerate(spans):
                cd = cells[j] if j < len(cells) else None
                if isinstance(cd, dict):
                    text = str(cd.get("text", ""))
                    align = cd.get("align") or "left"
                else:
                    text = "" if cd is None else str(cd)
                    align = "left"
                    cd = {}
                fmt = _cell_font_spec(cd)
                max_lines = max(max_lines, _cell_lines(text, caps[j]))
                # 点灯列：取值命中红/黄/绿则整格上底色 + 同色粗体，与页面和周报一致
                light = (enums.GRID_LIGHT_COLORS.get(text.strip())
                         if j < len(col_types) and col_types[j] == "light" else None)
                if c2 > c1:
                    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
                for cc in range(c1, c2 + 1):
                    cell = ws.cell(row=r, column=cc, value=text if cc == c1 else None)
                    cell.font = Font(name=fmt["name"], size=fmt["size"],
                                     color=_LIGHT_FONT[light] if light else fmt["color"],
                                     bold=bool(light) or fmt["bold"],
                                     italic=fmt["italic"], underline=fmt["underline"])
                    cell.alignment = Alignment(horizontal="center" if light else align,
                                               vertical="center", wrap_text=True)
                    cell.border = _BORDER
                    if light:
                        cell.fill = PatternFill("solid", fgColor=_LIGHT_FILL[light])
                    elif fmt["bg"]:
                        cell.fill = PatternFill("solid", fgColor=fmt["bg"])
                    elif zebra:
                        cell.fill = PatternFill("solid", fgColor=_ZEBRA)
            ws.row_dimensions[r].height = max(18, min(180, max_lines * 15 + 3))
            state["row"] += 1
        return True

    def render_block_images(block):
        items = [i for i in (block.get("items") or [])
                 if isinstance(i, dict) and i.get("file")]
        if not items:
            return False
        _flush_section()
        drawn = 0
        for im in items:
            if _place_image(ws, state, UPLOAD_ROOT / str(special.id) / str(im["file"])):
                drawn += 1
        if items and not drawn:
            # SVG 等 openpyxl 塞不进去的格式：说清楚，别让人以为图丢了
            narrative(f"（{len(items)} 张图片，Excel 无法内嵌，请见系统页面）")
        return bool(items)

    # ===== 各分段按「版式」顺序输出 =====
    # 标题与顺序全部来自 special_layout。**一个专项导出成一张表**：自定义表格以前
    # 因为列宽对不上被丢到独立工作表、主表只留一行"→ 见工作表「X」"，一份周报打开
    # 是四五个页签，而顺序、章节号、上下文全断在那一行指引上。现在物理列改成 36 个
    # 等宽窄列，自由表格按比例合并进来，所有分段回到同一张表上按顺序往下排。
    n = 0
    for sec in special_layout.resolve_sections(special):
        n += 1
        title = f"{_cn_index(n)}、{sec.title}"
        # 章节标题行还没落笔（见 _flush_section），但它就落在这一行
        sec_top = state["row"]
        section(title)
        ok = True
        if sec.is_custom and sec.kind == "grid":
            ok = render_grid(sec.block)
        elif sec.is_custom and sec.kind == "text":
            body = _strip_html(sec.block.get("html") or "")
            ok = bool(body.strip())
            if ok:
                narrative(body)
        elif sec.is_custom and sec.kind == "milestones":
            ok = render_milestones(sec.block.get("milestones") or [])
        elif sec.is_custom:
            ok = render_block_images(sec.block)
        elif sec.key in _TEXT_FIELD:
            body = _strip_html(getattr(content, _TEXT_FIELD[sec.key], "") or "") if content else ""
            ok = bool(body.strip())
            if ok:
                narrative(body)
        elif sec.key == "plan":
            ok = render_milestones()
        elif sec.key == "panorama":
            path = getattr(content, "panorama_image_path", "") if content else ""
            ok = bool(path)
            if ok:
                _flush_section()
            if ok and not _place_image(ws, state, UPLOAD_ROOT.parent / path):
                narrative(f"（{getattr(content, 'panorama_image_name', '') or '全景图'}："
                          f"Excel 无法内嵌该格式，请见系统页面）")
        elif sec.key == "risks":
            ok = render_items(special.risks or [], "问题内容")
        elif sec.key == "tasks":
            ok = render_items(sorted(special.tasks or [],
                                     key=lambda t: (t.sort_order or 0, t.id)), "事务内容")
        elif sec.key == "formation":
            ok = render_formation()
        else:
            ok = False
        if not ok:
            # 标题还挂着没落笔，撤回即可；编号让给下一段，与周报保持同一套章节号
            state["pending"] = None
            n -= 1
            continue
        # **整个分段（标题行 + 内容）压一个外框**：章节标题、正文、表格三层里
        # 只有表格自带细网格，没有外框时整页看着是一堆浮着的横条。
        _frame(ws, sec_top, state["row"] - 1)
        gap()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# 内置文本类分段 -> content 上的列（与 routers/specials.py 的同名表一致）
_TEXT_FIELD = {"goal": "goal", "progress": "progress_summary", "help": "help_request"}
_CN_DIGITS = "零一二三四五六七八九"


def _cn_index(n: int) -> str:
    if n <= 0:
        return str(n)
    if n < 10:
        return _CN_DIGITS[n]
    if n < 20:
        return "十" + (_CN_DIGITS[n - 10] if n > 10 else "")
    tens, ones = divmod(n, 10)
    return _CN_DIGITS[tens] + "十" + (_CN_DIGITS[ones] if ones else "")


def _cell_str(c) -> str:
    if isinstance(c, dict):
        return str(c.get("text") or "")
    return "" if c is None else str(c)


def _grid_cells(block: dict):
    """自定义表格 → (表头名, 非空数据行)。判断"这段有没有内容"用。"""
    headers = []
    for h in (block.get("headers") or []):
        if isinstance(h, dict):
            headers.extend([str(h.get("text") or "")] * max(1, int(h.get("colspan") or 1)))
        else:
            headers.append(str(h or ""))
    rows = [[_cell_str(c) for c in (r or [])] for r in (block.get("rows") or [])]
    return headers, [r for r in rows if any(x.strip() for x in r)]


def _place_image(ws, state, path) -> bool:
    """把磁盘上的图片放到当前行并预留行高；放不进去（SVG/缺 PIL/文件丢失）返回 False。"""
    try:
        p = pathlib.Path(path).resolve()
        if not p.exists() or p.suffix.lower() == ".svg":
            return False
        img = XLImage(str(p))
    except Exception:
        return False
    try:
        max_w = float(_SHEET_PX)
        if img.width and img.width > max_w:
            ratio = max_w / float(img.width)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
        r = state["row"]
        ws.add_image(img, f"A{r}")
        state["row"] += _rows_for_px(img.height or 200)
        return True
    except Exception:
        return False


def _cell_font_spec(cell: dict) -> dict:
    """单元格格式 → openpyxl 的 Font/Fill 参数。

    与 routers/specials._fmt_css、前端 utils/gridFormat.js 是同一张表的三种出口：
    页面要 CSS、周报要 CSS、Excel 要字体名 + 磅值。白名单外的值退回默认，
    因为 openpyxl 对非法字号会直接抛，而一次导出失败比丢个字号严重得多。
    """
    font_key = str(cell.get("font") or "")
    name = enums.GRID_FONTS.get(font_key, {}).get("xlsx") or _FONT
    try:
        px = int(cell.get("size") or 0)
    except (TypeError, ValueError):
        px = 0
    # px → 磅：Excel 用磅（1px ≈ 0.75pt）。默认 10 磅是本表原有的正文字号
    size = round(px * 0.75, 1) if px in enums.GRID_FONT_SIZES else 10
    bg = str(cell.get("bg") or "")
    return {
        "name": name,
        "size": size,
        "color": _hex_to_rgb6(cell.get("color", "")),
        "bold": bool(cell.get("bold")),
        "italic": bool(cell.get("italic")),
        "underline": "single" if cell.get("underline") else None,
        "bg": _hex_to_rgb6(bg, "") if bg in enums.GRID_CELL_BG and bg else "",
    }
